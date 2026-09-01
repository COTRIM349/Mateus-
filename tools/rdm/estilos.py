# -*- coding: utf-8 -*-
"""Padrao visual unico da planilha (fonte Arial em toda a pasta)."""
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

FONTE = "Arial"

AZUL_ESCURO = "1F3864"
AZUL_MEDIO = "2E5C8A"
CINZA_CLARO = "F2F2F2"
CINZA_BORDA = "BFBFBF"
AMARELO_INPUT = "FFF2CC"
VERMELHO_PEND = "FFC7CE"
VERDE_OK = "C6EFCE"
LARANJA = "FCE4D6"

TXT_INPUT = "0000FF"      # entrada digitada
TXT_FORMULA = "000000"    # calculado
TXT_LINK = "008000"       # vem de outra aba
TXT_PENDENTE = "C00000"   # dado pendente

_fina = Side(style="thin", color=CINZA_BORDA)
BORDA = Border(left=_fina, right=_fina, top=_fina, bottom=_fina)


def registrar(wb):
    """Cria os estilos nomeados usados por toda a pasta."""
    def novo(nome, **kw):
        if nome in wb.named_styles:
            return
        st = NamedStyle(name=nome)
        st.font = kw.get("font", Font(name=FONTE, size=10))
        st.alignment = kw.get("alignment", Alignment(vertical="center"))
        if "fill" in kw:
            st.fill = kw["fill"]
        if kw.get("borda", True):
            st.border = BORDA
        if "fmt" in kw:
            st.number_format = kw["fmt"]
        wb.add_named_style(st)

    novo("tituloPagina",
         font=Font(name=FONTE, size=16, bold=True, color="FFFFFF"),
         fill=PatternFill("solid", fgColor=AZUL_ESCURO),
         alignment=Alignment(vertical="center", horizontal="left", indent=1), borda=False)
    novo("subtitulo",
         font=Font(name=FONTE, size=10, italic=True, color="595959"),
         alignment=Alignment(vertical="center", horizontal="left", wrap_text=True), borda=False)
    novo("cabecalho",
         font=Font(name=FONTE, size=10, bold=True, color="FFFFFF"),
         fill=PatternFill("solid", fgColor=AZUL_MEDIO),
         alignment=Alignment(vertical="center", horizontal="center", wrap_text=True))
    novo("secao",
         font=Font(name=FONTE, size=11, bold=True, color=AZUL_ESCURO),
         alignment=Alignment(vertical="center"), borda=False)
    novo("rotulo", font=Font(name=FONTE, size=10, bold=True))
    novo("texto")
    novo("entrada", font=Font(name=FONTE, size=10, color=TXT_INPUT),
         fill=PatternFill("solid", fgColor=AMARELO_INPUT))
    novo("entradaNum", font=Font(name=FONTE, size=10, color=TXT_INPUT),
         fill=PatternFill("solid", fgColor=AMARELO_INPUT), fmt="#,##0.00")
    novo("entradaInt", font=Font(name=FONTE, size=10, color=TXT_INPUT),
         fill=PatternFill("solid", fgColor=AMARELO_INPUT), fmt="#,##0")
    novo("entradaPct", font=Font(name=FONTE, size=10, color=TXT_INPUT),
         fill=PatternFill("solid", fgColor=AMARELO_INPUT), fmt="0.0%")
    novo("entradaData", font=Font(name=FONTE, size=10, color=TXT_INPUT),
         fill=PatternFill("solid", fgColor=AMARELO_INPUT), fmt="dd/mm/yyyy")
    novo("num0", fmt="#,##0")
    novo("num1", fmt="#,##0.0")
    novo("num2", fmt="#,##0.00")
    novo("num3", fmt="#,##0.000")
    novo("pct", fmt="0.0%")
    novo("data", fmt="dd/mm/yyyy")
    novo("elo", font=Font(name=FONTE, size=10, color=TXT_LINK), fmt="#,##0.00")
    novo("eloTexto", font=Font(name=FONTE, size=10, color=TXT_LINK))
    novo("pendente", font=Font(name=FONTE, size=10, bold=True, color=TXT_PENDENTE),
         fill=PatternFill("solid", fgColor=VERMELHO_PEND),
         alignment=Alignment(vertical="center", horizontal="center"))
    novo("kpiRotulo", font=Font(name=FONTE, size=9, bold=True, color="FFFFFF"),
         fill=PatternFill("solid", fgColor=AZUL_MEDIO),
         alignment=Alignment(vertical="center", horizontal="center", wrap_text=True))
    novo("kpiValor", font=Font(name=FONTE, size=14, bold=True, color=AZUL_ESCURO),
         fill=PatternFill("solid", fgColor=CINZA_CLARO),
         alignment=Alignment(vertical="center", horizontal="center"), fmt="#,##0")
    novo("kpiValor1", font=Font(name=FONTE, size=14, bold=True, color=AZUL_ESCURO),
         fill=PatternFill("solid", fgColor=CINZA_CLARO),
         alignment=Alignment(vertical="center", horizontal="center"), fmt="#,##0.0")
    novo("kpiValorPct", font=Font(name=FONTE, size=14, bold=True, color=AZUL_ESCURO),
         fill=PatternFill("solid", fgColor=CINZA_CLARO),
         alignment=Alignment(vertical="center", horizontal="center"), fmt="0.0%")
    novo("kpiValorTxt", font=Font(name=FONTE, size=13, bold=True, color=AZUL_ESCURO),
         fill=PatternFill("solid", fgColor=CINZA_CLARO),
         alignment=Alignment(vertical="center", horizontal="center"))
    novo("kpiValorData", font=Font(name=FONTE, size=13, bold=True, color=AZUL_ESCURO),
         fill=PatternFill("solid", fgColor=CINZA_CLARO),
         alignment=Alignment(vertical="center", horizontal="center"), fmt="dd/mm/yyyy")
    novo("aviso", font=Font(name=FONTE, size=10, bold=True, color="833C00"),
         fill=PatternFill("solid", fgColor=LARANJA),
         alignment=Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1))


def cabecalho_pagina(ws, titulo, subtitulo, ultima_coluna="N"):
    ws["B2"] = titulo
    ws["B2"].style = "tituloPagina"
    ws.merge_cells("B2:%s2" % ultima_coluna)
    ws.row_dimensions[2].height = 30
    ws["B3"] = subtitulo
    ws["B3"].style = "subtitulo"
    ws.merge_cells("B3:%s3" % ultima_coluna)
    ws.row_dimensions[3].height = 26
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
