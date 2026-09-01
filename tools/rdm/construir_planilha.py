# -*- coding: utf-8 -*-
"""Constroi o simulador de planejamento hidrico da RDM.

Uso: python3 construir_planilha.py rdm_dataset.json SAIDA.xlsx
"""
import json
import sys
import unicodedata
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation

import config_modelo as cfg
import estilos as est

MESES = ["", "Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho", "Julho",
         "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]


def sa(s):
    """Remove acentos e padroniza para uso como chave."""
    if s is None:
        return ""
    s = "".join(c for c in unicodedata.normalize("NFD", str(s))
                if unicodedata.category(c) != "Mn")
    return s.strip().upper()


def dt(s):
    return datetime.strptime(s, "%Y-%m-%d").date() if s else None


# ===================================================================== dados
def preparar(caminho_json):
    d = json.load(open(caminho_json, encoding="utf-8"))
    ini = dt(cfg.SEMANA_1_SEGUNDA)
    fim = ini + timedelta(days=cfg.N_SEMANAS * 7 - 1)

    # ---- semanas
    semanas = []
    for w in range(1, cfg.N_SEMANAS + 1):
        a = ini + timedelta(days=(w - 1) * 7)
        semanas.append({"n": w, "ini": a, "fim": a + timedelta(days=6),
                        "meio": a + timedelta(days=3)})

    # ---- duracao de referencia por cultivar (arquivo de rotacao)
    dur_cultivar = {}
    for c in d["cultivares"]:
        if c["duracao_dias"]:
            dur_cultivar[(sa(c["cultura"]), sa(c["cultivar"]))] = int(c["duracao_dias"])

    # ---- ciclos: normaliza, estima plantio de C0, classifica grupo
    ciclos, pendencias = [], []
    for c in d["ciclos"]:
        cultura = sa(c["cultura"])
        cultivar = c["cultivar"] or ""
        pivo, fim_c = c["pivo"], dt(c["fim"])
        plantio = dt(c["plantio"])
        if not cultura or not fim_c:
            if c["cultura"] or c["cultivar"] or c["plantio"] or c["fim"]:
                pendencias.append({
                    "variavel": "Ciclo incompleto na rotacao", "equipamento": pivo,
                    "dado": "%s: cultura=%s cultivar=%s plantio=%s fim=%s"
                            % (c["rotulo"], c["cultura"], c["cultivar"], c["plantio"], c["fim"]),
                    "unidade": "-",
                    "impacto": "Ciclo NAO entra no motor de demanda; o pivo fica sem cultura nessas semanas."})
            continue
        if fim_c < ini or (plantio and plantio > fim):
            continue                                   # fora do horizonte da safra
        origem = "Rotacao (informado)"
        if plantio is None:                            # cultura anterior: so ha data-fim
            dur = dur_cultivar.get((cultura, sa(cultivar))) or cfg.DURACAO_REF.get(cultura, 120)
            plantio = fim_c - timedelta(days=dur)
            origem = "ESTIMADO = Data_Fim - duracao de referencia (%d d)" % dur
            pendencias.append({
                "variavel": "Data de plantio do ciclo anterior", "equipamento": pivo,
                "dado": "%s %s - encerramento em %s" % (cultura, cultivar, fim_c.strftime("%d/%m/%Y")),
                "unidade": "data",
                "impacto": "Kc do residuo de ciclo estimado por retro-calculo; confirmar a data real de plantio."})
        dias = (fim_c - plantio).days
        if dias <= 0:
            continue
        ciclos.append({"pivo": pivo, "cultura": cultura, "cultivar": cultivar,
                       "plantio": plantio, "fim": fim_c, "dias": dias,
                       "grupo": cfg.grupo_por_duracao(cultura, dias),
                       "ordem": c["ordem"], "origem": origem,
                       "area_rotacao": c["area_rotacao_ha"]})

    ciclos.sort(key=lambda x: (int(x["pivo"].split()[1]), x["plantio"]))

    # ---- sobreposicao de ciclos no mesmo pivo (o motor usa 1 ciclo por semana)
    por_pivo = {}
    for c in ciclos:
        por_pivo.setdefault(c["pivo"], []).append(c)
    for pivo, lst in por_pivo.items():
        for a, b in zip(lst, lst[1:]):
            if b["plantio"] <= a["fim"]:
                c_ini, c_fim = b["plantio"], min(a["fim"], b["fim"])
                pendencias.append({
                    "variavel": "Sobreposicao de ciclos", "equipamento": pivo,
                    "dado": "%s (ate %s) x %s (a partir de %s)"
                            % (a["cultura"], a["fim"].strftime("%d/%m/%Y"),
                               b["cultura"], b["plantio"].strftime("%d/%m/%Y")),
                    "unidade": "%d dias" % ((c_fim - c_ini).days + 1),
                    "impacto": "Nas semanas sobrepostas o modelo soma dois IDs de ciclo e nao encontra cultura. Ajustar as datas na rotacao."})

    # ---- pivos: cadastro oficial + area da rotacao
    area_rot = {c["pivo"]: c["area_rotacao"] for c in d["ciclos"] if c.get("area_rotacao")}
    pivos = []
    for p in d["pivos"]:
        pivos.append(dict(p, area_rotacao_ha=area_rot.get(p["pivo"])))
    pivos.sort(key=lambda x: (x["casa"], int(x["pivo"].split()[1])))

    casas = sorted(d["casas"], key=lambda x: x["casa"])
    for c in casas:
        c["pivos"] = [p["pivo"] for p in pivos if p["casa"] == c["casa"]]
    return {"semanas": semanas, "ciclos": ciclos, "pivos": pivos, "casas": casas,
            "pocos": d["pocos"], "reservacao": d["reservacao"],
            "cultivares": d["cultivares"], "pendencias": pendencias,
            "ini": ini, "fim": fim}


CAB, DAT = 5, 6          # linha de cabecalho / primeira linha de dados


def escrever_tabela(ws, cabecalhos, larguras, linha_cab=CAB, col_ini=2):
    for i, (h, w) in enumerate(zip(cabecalhos, larguras)):
        c = ws.cell(row=linha_cab, column=col_ini + i, value=h)
        c.style = "cabecalho"
        ws.column_dimensions[gcl(col_ini + i)].width = w
    ws.row_dimensions[linha_cab].height = 30
    ws.freeze_panes = ws.cell(row=linha_cab + 1, column=col_ini)


def nota(ws, linha, texto, ultima_coluna, estilo="aviso"):
    c = ws.cell(row=linha, column=2, value=texto)
    c.style = estilo
    ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=ultima_coluna)
    ws.row_dimensions[linha].height = 30


# ============================================================ PARAMETROS_GERAIS
def aba_parametros(wb, D):
    ws = wb.create_sheet("PARAMETROS_GERAIS")
    est.cabecalho_pagina(ws, "PARAMETROS GERAIS DO MODELO",
                         "Toda constante do modelo mora aqui. Nenhuma formula da pasta traz numero "
                         "digitado: as celulas amarelas sao as unicas entradas.", "G")
    escrever_tabela(ws, ["Nome definido", "Parametro", "Valor", "Unidade", "Origem / observacao", "Editavel"],
                    [30, 44, 14, 14, 74, 11])
    linha = DAT
    mapa = {}
    for nome, rotulo, valor, unidade, origem, editavel in cfg.PARAMETROS:
        ws.cell(row=linha, column=2, value=nome).style = "rotulo"
        ws.cell(row=linha, column=3, value=rotulo).style = "texto"
        c = ws.cell(row=linha, column=4)
        if nome == "HORAS_OPERACAO_SEMANA":
            c.value = "=HORAS_OPERACAO_DIA*DIAS_POR_SEMANA"
            c.style = "num0"
        else:
            c.value = valor
            c.style = ("entrada" if isinstance(valor, str) else
                       "entradaPct" if isinstance(valor, float) and valor <= 1 and unidade == "fracao"
                       else "entradaNum") if editavel else "num2"
        ws.cell(row=linha, column=5, value=unidade).style = "texto"
        o = ws.cell(row=linha, column=6, value=origem)
        o.style = "pendente" if origem.startswith("DADO PENDENTE") else "texto"
        ws.cell(row=linha, column=7, value="SIM" if editavel else "calculado").style = "texto"
        mapa[nome] = "PARAMETROS_GERAIS!$D$%d" % linha
        linha += 1

    # datas do horizonte
    linha += 1
    ws.cell(row=linha, column=2, value="HORIZONTE DA SAFRA").style = "secao"
    linha += 1
    for nome, rotulo, valor, fmt in (
            ("DATA_INICIO_SAFRA", "Inicio da semana 1 (segunda-feira)", D["ini"], "entradaData"),
            ("N_SEMANAS", "Numero de semanas simuladas", cfg.N_SEMANAS, "num0"),
            ("DATA_FIM_SAFRA", "Fim da ultima semana", D["fim"], "data"),
            ("SAFRA", "Identificacao da safra", cfg.SAFRA, "entrada")):
        ws.cell(row=linha, column=2, value=nome).style = "rotulo"
        ws.cell(row=linha, column=3, value=rotulo).style = "texto"
        ws.cell(row=linha, column=4, value=valor).style = fmt
        mapa[nome] = "PARAMETROS_GERAIS!$D$%d" % linha
        linha += 1

    dv = DataValidation(type="list", formula1='"REFERENCIA,USUARIO"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(mapa["CENARIO_CLIMA"].replace("PARAMETROS_GERAIS!", ""))

    nota(ws, linha + 1,
         "LEGENDA  -  celula AMARELA com texto AZUL = entrada do usuario  |  texto PRETO = formula  |  "
         "texto VERDE = valor trazido de outra aba  |  faixa VERMELHA = DADO PENDENTE, valor provisorio.", 7)
    for nome, ref in mapa.items():
        wb.defined_names.add(_dn(nome, ref))
    return mapa


def _dn(nome, ref):
    from openpyxl.workbook.defined_name import DefinedName
    return DefinedName(nome, attr_text=ref)


# ============================================================== CADASTRO_PIVOS
def aba_pivos(wb, D):
    ws = wb.create_sheet("CADASTRO_PIVOS")
    est.cabecalho_pagina(ws, "CADASTRO_PIVOS  -  hidraulica por pivo",
                         "Area e casa de bomba: arquivo oficial RDM (aba Parametros). Lamina nominal: maior "
                         "lamina praticada por pivo no arquivo oficial (aba data_Base). Vazao: preencha a coluna "
                         "de cadastro para sobrepor o calculo.", "P")
    escrever_tabela(ws, ["Pivo", "Casa de bomba", "Trecho de canal", "Area (ha)",
                         "Area rotacao (ha)", "Lamina nominal (mm/dia)", "Vazao cadastrada (m3/h)",
                         "Vazao do pivo (m3/h)", "Origem da vazao", "Eficiencia de aplicacao",
                         "Horas max/dia", "Horas max/semana", "Status", "Observacao"],
                    [11, 14, 14, 11, 13, 14, 14, 14, 30, 13, 12, 13, 13, 46])
    for i, p in enumerate(D["pivos"]):
        r = DAT + i
        ws.cell(row=r, column=2, value=p["pivo"]).style = "rotulo"
        ws.cell(row=r, column=3, value=p["casa"]).style = "texto"
        ws.cell(row=r, column=4,
                value="=IFERROR(INDEX(CADASTRO_CASAS_BOMBA!$C$%d:$C$%d,MATCH($C%d,CADASTRO_CASAS_BOMBA!$B$%d:$B$%d,0)),\"PENDENTE\")"
                      % (DAT, DAT + 6, r, DAT, DAT + 6)).style = "eloTexto"
        ws.cell(row=r, column=5, value=p["area_ha"]).style = "num2"
        ws.cell(row=r, column=6, value=p.get("area_rotacao_ha")).style = "num2"
        ws.cell(row=r, column=7, value=p.get("lamina_max_mm_dia")).style = "entradaNum"
        ws.cell(row=r, column=8).style = "entradaNum"          # vazao cadastrada (vazia)
        ws.cell(row=r, column=9,
                value="=IF(N($H%d)>0,$H%d,IF(N($G%d)>0,$E%d*$G%d*MM_PARA_M3_HA/HORAS_OPERACAO_DIA,0))"
                      % (r, r, r, r, r)).style = "num1"
        ws.cell(row=r, column=10,
                value='=IF(N($H%d)>0,"Cadastro do usuario","Calculada: Area x Lamina nominal x 10 / %s")'
                      % (r, "21 h")).style = "texto"
        ws.cell(row=r, column=11, value="=EFICIENCIA_PADRAO").style = "pct"
        ws.cell(row=r, column=12, value="=HORAS_OPERACAO_DIA").style = "num0"
        ws.cell(row=r, column=13, value="=HORAS_OPERACAO_SEMANA").style = "num0"
        ws.cell(row=r, column=14,
                value='=IF(N($I%d)<=0,"DADO PENDENTE","ATIVO")' % r).style = "texto"
        ws.cell(row=r, column=15,
                value="Lamina nominal = maior lamina/dia registrada no arquivo oficial (%d dias com registro)."
                      % (p.get("dias_com_lamina") or 0)).style = "texto"
    fim = DAT + len(D["pivos"]) - 1
    ws.conditional_formatting.add("N%d:N%d" % (DAT, fim),
                                  CellIsRule(operator="equal", formula=['"DADO PENDENTE"'],
                                             fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))
    nota(ws, fim + 2,
         "Para usar a vazao real de projeto de um pivo, digite-a em 'Vazao cadastrada (m3/h)'. "
         "Enquanto estiver vazia, o modelo reproduz a regra do arquivo oficial da RDM: "
         "Vazao = Area x Lamina nominal x 10 / horas de operacao por dia.", 15)
    return {"ini": DAT, "fim": fim}


# ======================================================= CADASTRO_CASAS_BOMBA
def aba_casas(wb, D, RP):
    ws = wb.create_sheet("CADASTRO_CASAS_BOMBA")
    est.cabecalho_pagina(ws, "CADASTRO_CASAS_BOMBA",
                         "Pocos, vazao de captacao e reservacao vem do arquivo oficial RDM (aba Dashboard). "
                         "A vazao OPERACIONAL de recalque de cada casa nao consta nos arquivos: preencha a coluna "
                         "amarela. Enquanto vazia, o modelo usa a soma das vazoes dos pivos da casa.", "R")
    escrever_tabela(ws, ["Casa de bomba", "Trecho de canal", "Pivos atendidos", "Area total atendida (ha)",
                         "Vazao nominal (m3/h)", "Vazao operacional cadastrada (m3/h)",
                         "Vazao operacional usada (m3/h)", "Vazao captacao pocos (m3/h)",
                         "Volume captacao (m3/dia)", "Horas disp./dia", "Horas disp./semana",
                         "Volume max/dia (m3)", "Volume max/semana (m3)", "Reservacao total (m3)",
                         "Pocos", "Observacao"],
                    [14, 13, 34, 15, 14, 16, 16, 15, 15, 12, 13, 14, 15, 15, 22, 40])
    trechos = []
    for i, c in enumerate(D["casas"]):
        r = DAT + i
        trecho = "TR%02d" % (i + 1)
        trechos.append(trecho)
        ws.cell(row=r, column=2, value=c["casa"]).style = "rotulo"
        ws.cell(row=r, column=3, value=trecho).style = "entrada"
        ws.cell(row=r, column=4, value=", ".join(c["pivos"])).style = "texto"
        ws.cell(row=r, column=5, value="=SUMIF(CADASTRO_PIVOS!$C$%d:$C$%d,$B%d,CADASTRO_PIVOS!$E$%d:$E$%d)"
                % (RP["ini"], RP["fim"], r, RP["ini"], RP["fim"])).style = "num2"
        ws.cell(row=r, column=6, value="=SUMIF(CADASTRO_PIVOS!$C$%d:$C$%d,$B%d,CADASTRO_PIVOS!$I$%d:$I$%d)"
                % (RP["ini"], RP["fim"], r, RP["ini"], RP["fim"])).style = "num1"
        ws.cell(row=r, column=7).style = "entradaNum"
        ws.cell(row=r, column=8, value="=IF(N($G%d)>0,$G%d,$F%d)" % (r, r, r)).style = "num1"
        ws.cell(row=r, column=9, value=c["vazao_captacao_m3h"]).style = "num1"
        ws.cell(row=r, column=10, value=c["volume_captacao_m3_dia"]).style = "num0"
        ws.cell(row=r, column=11, value="=HORAS_OPERACAO_DIA").style = "num0"
        ws.cell(row=r, column=12, value="=HORAS_OPERACAO_SEMANA").style = "num0"
        ws.cell(row=r, column=13, value="=$H%d*$K%d" % (r, r)).style = "num0"
        ws.cell(row=r, column=14, value="=$H%d*$L%d" % (r, r)).style = "num0"
        ws.cell(row=r, column=15, value=c["capacidade_total_m3"]).style = "num0"
        ws.cell(row=r, column=16, value=", ".join(c["pocos"])).style = "texto"
        ws.cell(row=r, column=17,
                value='=IF(N($G%d)>0,"Vazao operacional cadastrada pelo usuario.",'
                      '"PENDENTE: usando a soma das vazoes dos pivos como vazao operacional.")' % r).style = "texto"
    fim = DAT + len(D["casas"]) - 1
    ws.conditional_formatting.add("Q%d:Q%d" % (DAT, fim),
                                  CellIsRule(operator="beginsWith", formula=['"PENDENTE"'],
                                             fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))
    nota(ws, fim + 2,
         "Volume maximo/semana = Vazao operacional x HORAS_OPERACAO_DIA x DIAS_POR_SEMANA "
         "(21 x 7 = 147 h por padrao, alteravel em PARAMETROS_GERAIS).", 17)
    return {"ini": DAT, "fim": fim, "trechos": trechos}


# ==================================================== CADASTRO_TRECHOS_CANAL
def aba_trechos(wb, D, RC):
    ws = wb.create_sheet("CADASTRO_TRECHOS_CANAL")
    est.cabecalho_pagina(ws, "CADASTRO_TRECHOS_CANAL  -  topologia hidraulica do canal",
                         "A topologia do canal NAO consta nos arquivos oficiais. A matriz TRECHO x CASA abaixo "
                         "e a unica definicao da rede: marque 1 onde o trecho transporta a agua daquela casa. "
                         "Em canal em serie, um trecho a montante deve receber 1 em todas as casas a jusante.", "N")
    escrever_tabela(ws, ["ID_Trecho", "Nome_Trecho", "Trecho_Inicial", "Trecho_Final",
                         "Casa_Inicial", "Casa_Final", "Casas_Atendidas",
                         "Vazao_Maxima_m3h", "Vazao_Operacional_m3h", "Vazao usada (m3/h)",
                         "Volume_Maximo_Dia_m3", "Volume canal armazenado (m3)", "Observacao"],
                    [11, 22, 14, 14, 13, 13, 30, 15, 16, 15, 16, 18, 52])
    casas = D["casas"]
    n = len(casas)
    lin_mapa_cab = DAT + n + 3
    lin_mapa = lin_mapa_cab + 1
    for i, c in enumerate(casas):
        r = DAT + i
        t = RC["trechos"][i]
        ws.cell(row=r, column=2, value=t).style = "rotulo"
        ws.cell(row=r, column=3, value="Trecho %s (%s)" % (t[-2:], c["casa"])).style = "entrada"
        ws.cell(row=r, column=4).style = "entrada"
        ws.cell(row=r, column=5).style = "entrada"
        ws.cell(row=r, column=6).style = "entrada"
        ws.cell(row=r, column=7).style = "entrada"
        alvos = ",".join('IF(%s%d=1,%s$%d,"")' % (gcl(3 + j), lin_mapa + i, gcl(3 + j), lin_mapa_cab)
                         for j in range(n))
        ws.cell(row=r, column=8, value='=_xlfn.TEXTJOIN(", ",TRUE,%s)' % alvos).style = "eloTexto"
        ws.cell(row=r, column=9).style = "entradaNum"
        ws.cell(row=r, column=10).style = "entradaNum"
        # soma explicita termo a termo: a linha da matriz e horizontal e a coluna de
        # vazoes das casas e vertical, entao SOMARPRODUTO daria erro de dimensao.
        soma_casas = "+".join("$%s$%d*CADASTRO_CASAS_BOMBA!$H$%d"
                              % (gcl(3 + j), lin_mapa + i, RC["ini"] + j) for j in range(n))
        ws.cell(row=r, column=11,
                value="=IF(N($J%d)>0,$J%d,IF(N($I%d)>0,$I%d,%s))"
                      % (r, r, r, r, soma_casas)).style = "num1"
        ws.cell(row=r, column=12, value="=$K%d*HORAS_OPERACAO_DIA" % r).style = "num0"
        ws.cell(row=r, column=13, value=(D["reservacao"].get(c["casa"], {}) or {}).get("canal_m3")).style = "num0"
        ws.cell(row=r, column=14,
                value='=IF(N($J%d)>0,"Capacidade operacional cadastrada.",'
                      '"DADO PENDENTE: sem capacidade de canal cadastrada. O modelo assume que o trecho entrega '
                      'exatamente a soma das casas atendidas, ou seja, utilizacao 100%% e deficit zero por construcao.")'
                      % r).style = "texto"
    fim = DAT + n - 1
    ws.conditional_formatting.add("N%d:N%d" % (DAT, fim),
                                  CellIsRule(operator="beginsWith", formula=['"DADO PENDENTE"'],
                                             fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))

    # ---------------- matriz TRECHO x CASA
    ws.cell(row=lin_mapa_cab - 2, column=2,
            value="MATRIZ DE DEPENDENCIA  TRECHO x CASA DE BOMBA  (1 = o trecho transporta a agua da casa)"
            ).style = "secao"
    ws.cell(row=lin_mapa_cab, column=2, value="ID_Trecho").style = "cabecalho"
    for j, c in enumerate(casas):
        ws.cell(row=lin_mapa_cab, column=3 + j, value=c["casa"]).style = "cabecalho"
    for i, t in enumerate(RC["trechos"]):
        ws.cell(row=lin_mapa + i, column=2, value=t).style = "rotulo"
        for j in range(n):
            cel = ws.cell(row=lin_mapa + i, column=3 + j, value=1 if i == j else 0)
            cel.style = "entradaInt"
    ws.cell(row=lin_mapa + n, column=2, value="Casas por trecho").style = "rotulo"
    for j in range(n):
        ws.cell(row=lin_mapa + n, column=3 + j,
                value="=SUM(%s%d:%s%d)" % (gcl(3 + j), lin_mapa, gcl(3 + j), lin_mapa + n - 1)).style = "num0"
    nota(ws, lin_mapa + n + 2,
         "TOPOLOGIA PROVISORIA: a matriz vem preenchida como 1 trecho por casa (diagonal). Isso NAO representa "
         "um canal em serie. Assim que a topologia real for informada, marque 1 em todas as casas a jusante de "
         "cada trecho e todos os calculos de canal se ajustam sozinhos.", 14)
    return {"ini": DAT, "fim": fim, "mapa_cab": lin_mapa_cab, "mapa": lin_mapa, "n": n}


# ====================================================== PARAMETROS_CULTURAS
def aba_culturas(wb, D, RCL):
    ws = wb.create_sheet("PARAMETROS_CULTURAS")
    est.cabecalho_pagina(ws, "PARAMETROS_CULTURAS  -  curva de Kc por cultura e grupo de ciclo",
                         "Kc NAO e constante: a curva e definida por 4 fases (inicial, desenvolvimento, media, final) "
                         "expressas como fracao do ciclo, com interpolacao linear entre os pontos. "
                         "Fonte dos valores de partida: %s" % cfg.FONTE_KC, "Q")
    escrever_tabela(ws, ["Chave (Cultura|Grupo)", "Cultura", "Grupo de ciclo", "Duracao referencia (dias)",
                         "Fase inicial (fracao)", "Fase desenvolvimento (fracao)", "Fase media (fracao)",
                         "Fase final (fracao)", "Soma das fases", "Kc inicial", "Kc medio (fase media)",
                         "Kc final", "Kc medio ponderado do ciclo", "Consumo potencial (mm/ciclo)",
                         "Observacao"],
                    [24, 17, 14, 14, 13, 16, 13, 12, 12, 11, 14, 11, 16, 17, 46])
    for i, k in enumerate(cfg.CURVAS_KC):
        r = DAT + i
        chave, cultura, grupo, dur, fi, fd, fm, ff, ki, km, kf = k
        ws.cell(row=r, column=2, value=chave).style = "rotulo"
        ws.cell(row=r, column=3, value=cultura).style = "texto"
        ws.cell(row=r, column=4, value=grupo).style = "texto"
        ws.cell(row=r, column=5, value=dur).style = "entradaInt"
        for j, v in enumerate((fi, fd, fm, ff)):
            ws.cell(row=r, column=6 + j, value=v).style = "entradaPct"
        ws.cell(row=r, column=10, value="=SUM($F%d:$I%d)" % (r, r)).style = "pct"
        for j, v in enumerate((ki, km, kf)):
            ws.cell(row=r, column=11 + j, value=v).style = "entradaNum"
        ws.cell(row=r, column=14,
                value="=$F%d*$K%d+$G%d*($K%d+$L%d)/2+$H%d*$L%d+$I%d*($L%d+$M%d)/2"
                      % (r, r, r, r, r, r, r, r, r, r)).style = "num3"
        ws.cell(row=r, column=15,
                value="=$N%d*$E%d*AVERAGE(CLIMA_ETo_CHUVA!$G$%d:$G$%d)" % (r, r, RCL["ini"], RCL["fim"])).style = "num0"
        ws.cell(row=r, column=16, value="Curva FAO-56; ajuste livre das fases e dos Kc.").style = "texto"
    fim = DAT + len(cfg.CURVAS_KC) - 1
    ws.conditional_formatting.add("J%d:J%d" % (DAT, fim),
                                  CellIsRule(operator="notEqual", formula=["1"],
                                             fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))
    nota(ws, fim + 2,
         "A soma das quatro fases precisa dar 100%. A coluna 'Soma das fases' fica vermelha se nao der. "
         "'Consumo potencial' usa a ETo media do horizonte simulado, so como referencia agronomica.", 16)
    return {"ini": DAT, "fim": fim}


# ============================================================== ROTACAO_RDM
def aba_rotacao(wb, D, RP):
    ws = wb.create_sheet("ROTACAO_RDM")
    est.cabecalho_pagina(ws, "ROTACAO_RDM  -  ocupacao agricola dos pivos da RDM",
                         "Importada do arquivo de rotacao bienal 26-28, filtrando FAZENDA = RDM e tipo = Irrigado. "
                         "Karitel e demais unidades ficaram de fora. Altere data de plantio, data final ou cultura "
                         "e o planejamento inteiro recalcula.", "R")
    escrever_tabela(ws, ["ID", "Safra", "Pivo", "Area (ha)", "Cultura", "Cultivar", "Grupo de ciclo",
                         "Data plantio", "Data final", "Duracao (dias)", "Casa de bomba", "Trecho de canal",
                         "Chave da curva de Kc", "Origem da data de plantio", "Ordem na rotacao", "Status"],
                    [7, 15, 11, 11, 16, 24, 13, 13, 13, 12, 13, 13, 24, 46, 14, 26])
    for i, c in enumerate(D["ciclos"]):
        r = DAT + i
        ws.cell(row=r, column=2, value=i + 1).style = "num0"
        ws.cell(row=r, column=3, value="=SAFRA").style = "eloTexto"
        ws.cell(row=r, column=4, value=c["pivo"]).style = "rotulo"
        ws.cell(row=r, column=5,
                value="=IFERROR(INDEX(CADASTRO_PIVOS!$E$%d:$E$%d,MATCH($D%d,CADASTRO_PIVOS!$B$%d:$B$%d,0)),0)"
                      % (RP["ini"], RP["fim"], r, RP["ini"], RP["fim"])).style = "elo"
        ws.cell(row=r, column=6, value=c["cultura"]).style = "entrada"
        ws.cell(row=r, column=7, value=c["cultivar"]).style = "entrada"
        ws.cell(row=r, column=8, value=c["grupo"]).style = "entrada"
        ws.cell(row=r, column=9, value=c["plantio"]).style = "entradaData"
        ws.cell(row=r, column=10, value=c["fim"]).style = "entradaData"
        ws.cell(row=r, column=11, value="=MAX(1,$J%d-$I%d)" % (r, r)).style = "num0"
        ws.cell(row=r, column=12,
                value="=IFERROR(INDEX(CADASTRO_PIVOS!$C$%d:$C$%d,MATCH($D%d,CADASTRO_PIVOS!$B$%d:$B$%d,0)),\"?\")"
                      % (RP["ini"], RP["fim"], r, RP["ini"], RP["fim"])).style = "eloTexto"
        ws.cell(row=r, column=13,
                value="=IFERROR(INDEX(CADASTRO_PIVOS!$D$%d:$D$%d,MATCH($D%d,CADASTRO_PIVOS!$B$%d:$B$%d,0)),\"?\")"
                      % (RP["ini"], RP["fim"], r, RP["ini"], RP["fim"])).style = "eloTexto"
        ws.cell(row=r, column=14, value="=$F%d&\"|\"&$H%d" % (r, r)).style = "texto"
        ws.cell(row=r, column=15, value=c["origem"]).style = (
            "pendente" if c["origem"].startswith("ESTIMADO") else "texto")
        ws.cell(row=r, column=16, value=c["ordem"]).style = "texto"
        ws.cell(row=r, column=17,
                value='=IF(ISNA(MATCH($N%d,PARAMETROS_CULTURAS!$B$%d:$B$%d,0)),'
                      '"SEM CURVA DE Kc","OK")' % (r, DAT, DAT + len(cfg.CURVAS_KC) - 1)).style = "texto"
    fim = DAT + len(D["ciclos"]) - 1
    ws.conditional_formatting.add("Q%d:Q%d" % (DAT, fim),
                                  CellIsRule(operator="equal", formula=['"SEM CURVA DE Kc"'],
                                             fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))
    nota(ws, fim + 2,
         "'Ordem na rotacao': C0 = cultura anterior ainda ocupando o pivo no inicio do horizonte, "
         "C1/C2/C3 = ciclos 1, 2 e 3 da rotacao bienal. A area usada no calculo e sempre a do CADASTRO_PIVOS "
         "(arquivo oficial), nao a area arredondada da planilha de rotacao.", 17)
    return {"ini": DAT, "fim": fim}


# ========================================================= CLIMA_ETo_CHUVA
def aba_clima(wb, D):
    ws = wb.create_sheet("CLIMA_ETo_CHUVA")
    est.cabecalho_pagina(ws, "CLIMA_ETo_CHUVA  -  entrada climatica diaria do modelo",
                         "Cenario deterministico, sem Monte Carlo. Os arquivos oficiais da RDM NAO trazem ETo nem "
                         "chuva: a curva de referencia abaixo e PROVISORIA e deve ser trocada por serie de estacao "
                         "ou banco climatico. Preencha as colunas de usuario para sobrepor dia a dia.", "R")
    nota(ws, 4, "ATENCAO - DADO PENDENTE: ETo e chuva de referencia sao valores de partida, NAO sao dado oficial "
                "da fazenda. Substitua antes de usar o modelo para decisao operacional.", 18, "pendente")
    escrever_tabela(ws, ["Data", "Semana", "Mes", "ETo referencia (mm/dia)", "ETo usuario (mm/dia)",
                         "ETo mm/dia (usada)", "ETo acumulada na semana (mm)", "Chuva referencia (mm)",
                         "Chuva usuario (mm)", "Chuva bruta mm (usada)", "Chuva util calculada (mm)",
                         "Chuva util usuario (mm)", "Chuva util mm (usada)", "Fonte",
                         "Metodo da chuva util", "Observacao"],
                    [12, 9, 12, 15, 15, 14, 16, 14, 14, 14, 15, 15, 14, 44, 40, 26], linha_cab=7)
    dat = 8
    dias = cfg.N_SEMANAS * 7
    for i in range(dias):
        r = dat + i
        dia = D["ini"] + timedelta(days=i)
        sem = i // 7 + 1
        ws.cell(row=r, column=2, value=dia).style = "data"
        ws.cell(row=r, column=3, value=sem).style = "num0"
        ws.cell(row=r, column=4, value=MESES[dia.month]).style = "texto"
        ws.cell(row=r, column=5, value=cfg.ETO_REF_MM_DIA[dia.month]).style = "entradaNum"
        ws.cell(row=r, column=6).style = "entradaNum"
        ws.cell(row=r, column=7,
                value='=IF(N($F%d)>0,$F%d,IF(CENARIO_CLIMA="USUARIO",0,$E%d))' % (r, r, r)).style = "num2"
        ws.cell(row=r, column=8,
                value="=SUMIFS($G$%d:$G$%d,$C$%d:$C$%d,$C%d)" % (dat, dat + dias - 1, dat, dat + dias - 1, r)
                ).style = "num1"
        dias_mes = (date(dia.year + (dia.month == 12), dia.month % 12 + 1, 1) - date(dia.year, dia.month, 1)).days
        ws.cell(row=r, column=9, value=round(cfg.CHUVA_REF_MM_MES[dia.month] / dias_mes, 2)).style = "entradaNum"
        ws.cell(row=r, column=10).style = "entradaNum"
        ws.cell(row=r, column=11,
                value='=IF(N($J%d)>0,$J%d,IF(CENARIO_CLIMA="USUARIO",0,$I%d))' % (r, r, r)).style = "num2"
        ws.cell(row=r, column=12,
                value="=MIN($K%d*FATOR_CHUVA_UTIL,LIMITE_CHUVA_UTIL_DIA)" % r).style = "num2"
        ws.cell(row=r, column=13).style = "entradaNum"
        ws.cell(row=r, column=14, value="=IF(N($M%d)>0,$M%d,$L%d)" % (r, r, r)).style = "num2"
        ws.cell(row=r, column=15,
                value='=IF(N($F%d)+N($J%d)>0,"Serie informada pelo usuario","%s")' % (r, r, cfg.FONTE_CLIMA)
                ).style = "texto"
        ws.cell(row=r, column=16,
                value='=IF(N($M%d)>0,"Chuva util informada diretamente pelo usuario",'
                      '"MIN(Chuva bruta x FATOR_CHUVA_UTIL ; LIMITE_CHUVA_UTIL_DIA)")' % r).style = "texto"
    fim = dat + dias - 1
    return {"ini": dat, "fim": fim}


# ================================================================= SEMANAS
def aba_semanas(wb, D, RCL):
    ws = wb.create_sheet("SEMANAS")
    est.cabecalho_pagina(ws, "SEMANAS  -  indice temporal da safra",
                         "Uma linha por semana da safra. Agrega a entrada climatica diaria e define as horas "
                         "operacionais disponiveis de cada semana.", "L")
    escrever_tabela(ws, ["Semana", "Data inicial", "Data final", "Data de referencia (meio)", "Mes",
                         "Dias", "ETo da semana (mm)", "Chuva bruta da semana (mm)",
                         "Chuva util da semana (mm)", "Horas disponiveis na semana"],
                    [9, 13, 13, 19, 13, 8, 15, 16, 16, 17])
    ci, cf = RCL["ini"], RCL["fim"]
    for i, s in enumerate(D["semanas"]):
        r = DAT + i
        ws.cell(row=r, column=2, value=s["n"]).style = "num0"
        ws.cell(row=r, column=3, value="=DATA_INICIO_SAFRA+($B%d-1)*DIAS_POR_SEMANA" % r).style = "data"
        ws.cell(row=r, column=4, value="=$C%d+DIAS_POR_SEMANA-1" % r).style = "data"
        ws.cell(row=r, column=5, value="=$C%d+INT(DIAS_POR_SEMANA/2)" % r).style = "data"
        ws.cell(row=r, column=6, value="=TEXT($E%d,\"mmm/yy\")" % r).style = "texto"
        ws.cell(row=r, column=7, value="=COUNTIFS(CLIMA_ETo_CHUVA!$C$%d:$C$%d,$B%d)" % (ci, cf, r)).style = "num0"
        ws.cell(row=r, column=8,
                value="=SUMIFS(CLIMA_ETo_CHUVA!$G$%d:$G$%d,CLIMA_ETo_CHUVA!$C$%d:$C$%d,$B%d)"
                      % (ci, cf, ci, cf, r)).style = "num1"
        ws.cell(row=r, column=9,
                value="=SUMIFS(CLIMA_ETo_CHUVA!$K$%d:$K$%d,CLIMA_ETo_CHUVA!$C$%d:$C$%d,$B%d)"
                      % (ci, cf, ci, cf, r)).style = "num1"
        ws.cell(row=r, column=10,
                value="=SUMIFS(CLIMA_ETo_CHUVA!$N$%d:$N$%d,CLIMA_ETo_CHUVA!$C$%d:$C$%d,$B%d)"
                      % (ci, cf, ci, cf, r)).style = "num1"
        ws.cell(row=r, column=11, value="=$G%d*HORAS_OPERACAO_DIA" % r).style = "num0"
    return {"ini": DAT, "fim": DAT + len(D["semanas"]) - 1}


# =================================================== PLANEJAMENTO_SEMANAL
COLS_PS = ["Semana", "Data inicial", "Data final", "Pivo", "Casa de bomba", "Trecho de canal",
           "Area (ha)", "ID do ciclo", "Linha da rotacao", "Cultura", "Cultivar", "Data de plantio",
           "DAE (dias)", "Fracao do ciclo", "Linha da curva Kc", "Fase / estadio", "Kc",
           "ETo da semana (mm)", "ETc (mm)", "Chuva estimada (mm)", "Chuva util (mm)",
           "Necessidade liquida (mm)", "Eficiencia", "Lamina bruta (mm)", "Volume (m3)",
           "Vazao do pivo (m3/h)", "Horas necessarias", "Horas disponiveis",
           "Utilizacao do pivo (%)", "Status", "Chave semana|utilizacao",
           "ETc x area (mm.ha)", "Chuva util x area (mm.ha)", "Nec. liquida x area (mm.ha)",
           "ETo x area (mm.ha)", "Chave pivo|utilizacao"]


def aba_planejamento(wb, D, RP, RR, RPC, RS):
    ws = wb.create_sheet("PLANEJAMENTO_SEMANAL")
    est.cabecalho_pagina(ws, "PLANEJAMENTO_SEMANAL  -  motor de calculo (uma linha por SEMANA x PIVO)",
                         "CLIMA -> ETo -> Kc -> ETc -> chuva util -> necessidade liquida -> lamina bruta -> volume -> "
                         "horas -> utilizacao. Toda a analise da pasta le esta tabela por SOMASES.", "AK")
    escrever_tabela(ws, COLS_PS,
                    [8, 12, 12, 10, 13, 13, 10, 10, 12, 16, 22, 13, 10, 12, 13, 17, 8, 13, 11,
                     14, 13, 15, 11, 13, 13, 14, 13, 13, 14, 14, 18, 15, 17, 18, 15, 18])
    ri, rf = RR["ini"], RR["fim"]
    pi, pf = RPC["ini"], RPC["fim"]
    si, sf = RS["ini"], RS["fim"]
    vi, vf = RP["ini"], RP["fim"]

    def rot(col):
        return "ROTACAO_RDM!$%s$%d:$%s$%d" % (col, ri, col, rf)

    def pc(col):
        return "PARAMETROS_CULTURAS!$%s$%d:$%s$%d" % (col, pi, col, pf)

    def sem(col):
        return "SEMANAS!$%s$%d:$%s$%d" % (col, si, col, sf)

    def piv(col):
        return "CADASTRO_PIVOS!$%s$%d:$%s$%d" % (col, vi, col, vf)

    r = DAT
    for p in D["pivos"]:
        for s in D["semanas"]:
            mp = "MATCH($E%d,%s,0)" % (r, piv("B"))
            ms = "MATCH($B%d,%s,0)" % (r, sem("B"))
            fi = "INDEX(%s,$P%d)" % (pc("F"), r)
            fd = "INDEX(%s,$P%d)" % (pc("G"), r)
            fm = "INDEX(%s,$P%d)" % (pc("H"), r)
            ff = "INDEX(%s,$P%d)" % (pc("I"), r)
            ki = "INDEX(%s,$P%d)" % (pc("K"), r)
            km = "INDEX(%s,$P%d)" % (pc("L"), r)
            kf = "INDEX(%s,$P%d)" % (pc("M"), r)
            meio = "$C%d+INT(DIAS_POR_SEMANA/2)" % r
            vals = [
                s["n"],
                "=DATA_INICIO_SAFRA+($B%d-1)*DIAS_POR_SEMANA" % r,
                "=$C%d+DIAS_POR_SEMANA-1" % r,
                p["pivo"],
                "=IFERROR(INDEX(%s,%s),\"?\")" % (piv("C"), mp),
                "=IFERROR(INDEX(%s,%s),\"?\")" % (piv("D"), mp),
                "=IFERROR(INDEX(%s,%s),0)" % (piv("E"), mp),
                "=IFERROR(SUMIFS(%s,%s,$E%d,%s,\"<=\"&%s,%s,\">=\"&%s),0)"
                % (rot("B"), rot("D"), r, rot("I"), meio, rot("J"), meio),
                "=IFERROR(MATCH($I%d,%s,0),0)" % (r, rot("B")),
                "=IF($J%d=0,\"SEM CULTURA\",INDEX(%s,$J%d))" % (r, rot("F"), r),
                "=IF($J%d=0,\"\",INDEX(%s,$J%d))" % (r, rot("G"), r),
                "=IF($J%d=0,\"\",INDEX(%s,$J%d))" % (r, rot("I"), r),
                "=IF($J%d=0,0,MAX(0,%s-INDEX(%s,$J%d)))" % (r, meio, rot("I"), r),
                "=IF($J%d=0,0,MIN(1,$N%d/MAX(1,INDEX(%s,$J%d))))" % (r, r, rot("K"), r),
                "=IF($J%d=0,0,IFERROR(MATCH(INDEX(%s,$J%d),%s,0),0))" % (r, rot("N"), r, pc("B")),
                "=IF($P%d=0,\"-\",IF($O%d<=%s,\"INICIAL\",IF($O%d<=%s+%s,\"DESENVOLVIMENTO\","
                "IF($O%d<=%s+%s+%s,\"MEDIA\",\"FINAL\"))))"
                % (r, r, fi, r, fi, fd, r, fi, fd, fm),
                "=IF($P%d=0,0,IF($O%d<=%s,%s,IF($O%d<=%s+%s,%s+(%s-%s)*($O%d-%s)/MAX(%s,0.0001),"
                "IF($O%d<=%s+%s+%s,%s,%s+(%s-%s)*($O%d-%s-%s-%s)/MAX(%s,0.0001)))))"
                % (r, r, fi, ki, r, fi, fd, ki, km, ki, r, fi, fd,
                   r, fi, fd, fm, km, km, kf, km, r, fi, fd, fm, ff),
                "=IFERROR(INDEX(%s,%s),0)" % (sem("H"), ms),
                "=$R%d*$S%d" % (r, r),
                "=IFERROR(INDEX(%s,%s),0)" % (sem("I"), ms),
                "=IFERROR(INDEX(%s,%s),0)" % (sem("J"), ms),
                "=MAX(0,$T%d-$V%d)" % (r, r),
                "=IFERROR(INDEX(%s,%s),EFICIENCIA_PADRAO)" % (piv("K"), mp),
                "=IF($X%d<=0,0,$W%d/$X%d)" % (r, r, r),
                "=$Y%d*$H%d*MM_PARA_M3_HA" % (r, r),
                "=IFERROR(INDEX(%s,%s),0)" % (piv("I"), mp),
                "=IF($Z%d<=0,0,$Z%d/$AA%d)" % (r, r, r),
                "=IFERROR(INDEX(%s,%s),0)" % (sem("K"), ms),
                "=IF($AC%d<=0,0,$AB%d/$AC%d)" % (r, r, r),
                "=IF($K%d=\"SEM CULTURA\",\"SEM CULTURA\",IF($AA%d<=0,\"DADO PENDENTE\","
                "IF($AD%d>LIMITE_CRITICO,\"INVIAVEL\",IF($AD%d>LIMITE_ATENCAO,\"CRITICO\","
                "IF($AD%d>LIMITE_CONFORTAVEL,\"ATENCAO\",\"OK\")))))" % (r, r, r, r, r),
                "=$B%d&\"|\"&TEXT($AD%d,\"0.000000\")" % (r, r),
                # ponderacoes por area: zeradas quando o pivo esta sem cultura, para que
                # ETo, ETc e chuva util do balanco cubram sempre o mesmo periodo ocupado.
                '=IF($K%d="SEM CULTURA",0,$T%d*$H%d)' % (r, r, r),
                '=IF($K%d="SEM CULTURA",0,$V%d*$H%d)' % (r, r, r),
                '=IF($K%d="SEM CULTURA",0,$W%d*$H%d)' % (r, r, r),
                '=IF($K%d="SEM CULTURA",0,$S%d*$H%d)' % (r, r, r),
                '=$E%d&"|"&TEXT($AD%d,"0.000000")' % (r, r),
            ]
            estilos = ["num0", "data", "data", "rotulo", "texto", "texto", "num2", "num0", "num0",
                       "texto", "texto", "data", "num0", "pct", "num0", "texto", "num3", "num1",
                       "num1", "num1", "num1", "num1", "pct", "num1", "num0", "num1", "num1",
                       "num0", "pct", "texto", "texto", "num1", "num1", "num1", "num1", "texto"]
            for j, (v, stl) in enumerate(zip(vals, estilos)):
                ws.cell(row=r, column=2 + j, value=v).style = stl
            r += 1
    fim = r - 1
    faixa = "AE%d:AE%d" % (DAT, fim)
    for txt, cor in (("OK", est.VERDE_OK), ("ATENCAO", "FFEB9C"),
                     ("CRITICO", est.LARANJA), ("INVIAVEL", est.VERMELHO_PEND),
                     ("DADO PENDENTE", est.VERMELHO_PEND)):
        ws.conditional_formatting.add(faixa, CellIsRule(
            operator="equal", formula=['"%s"' % txt],
            fill=est.PatternFill("solid", fgColor=cor)))
    ws.column_dimensions["J"].hidden = True
    ws.column_dimensions["P"].hidden = True
    for c in ("AF", "AG", "AH", "AI", "AJ", "AK"):
        ws.column_dimensions[c].hidden = True
    ws.auto_filter.ref = "B%d:AK%d" % (CAB, fim)
    return {"ini": DAT, "fim": fim}


# ============================================================ AUX_MATRIZES
def aba_aux(wb, D, RPS, RS, RC, RT):
    """Camada de calculo: matrizes CASA x SEMANA e TRECHO x SEMANA."""
    ws = wb.create_sheet("AUX_MATRIZES")
    est.cabecalho_pagina(ws, "AUX_MATRIZES  -  camada de calculo (CASA x SEMANA e TRECHO x SEMANA)",
                         "Aba tecnica. Os resumos, o dashboard, os graficos e os mapas de calor leem daqui. "
                         "Nao ha nada para preencher nesta aba.", "H")
    nsem = len(D["semanas"])
    casas = [c["casa"] for c in D["casas"]]
    trechos = RT_nomes = RC["trechos"]
    n = len(casas)
    ci, cf = RPS["ini"], RPS["fim"]
    ps = lambda col: "PLANEJAMENTO_SEMANAL!$%s$%d:$%s$%d" % (col, ci, col, cf)   # noqa: E731
    sem = lambda col: "SEMANAS!$%s$%d:$%s$%d" % (col, RS["ini"], col, RS["fim"])  # noqa: E731
    cas = lambda col: "CADASTRO_CASAS_BOMBA!$%s$%d:$%s$%d" % (col, RC["ini"], col, RC["fim"])  # noqa: E731

    blocos = [
        ("VOL_CASA", "Volume necessario por casa de bomba (m3)", casas, "num0"),
        ("AREA_CASA", "Area ativa por casa de bomba (ha)", casas, "num1"),
        ("HORAS_CASA", "Horas necessarias por casa de bomba (h)", casas, "num1"),
        ("QREQ_CASA", "Vazao media requerida por casa (m3/h)", casas, "num1"),
        ("UTIL_CASA", "Utilizacao da casa de bomba (%)", casas, "pct"),
        ("DEF_CASA", "Deficit de vazao da casa (m3/h)", casas, "num1"),
        ("VOL_TRECHO", "Volume necessario por trecho de canal (m3)", trechos, "num0"),
        ("QREQ_TRECHO", "Vazao requerida por trecho (m3/h)", trechos, "num1"),
        ("QDISP_TRECHO", "Vazao disponivel por trecho (m3/h)", trechos, "num1"),
        ("UTIL_TRECHO", "Utilizacao do trecho (%)", trechos, "pct"),
        ("DEF_TRECHO", "Deficit de vazao do trecho (m3/h)", trechos, "num1"),
    ]
    pos, linha = {}, 6
    for chave, titulo, nomes, _ in blocos:
        pos[chave] = {"titulo": linha, "cab": linha + 1, "ini": linha + 2,
                      "fim": linha + 1 + len(nomes)}
        linha = pos[chave]["fim"] + 3

    totais = ["Volume total necessario (m3)", "Area ativa total (ha)", "ETo da semana (mm)",
              "ETc media ponderada (mm)", "Chuva util media ponderada (mm)",
              "Necessidade liquida media (mm)", "Lamina bruta media (mm)",
              "Horas necessarias totais (h)", "Utilizacao maxima de pivo (%)", "Pivo mais critico",
              "Utilizacao da casa mais critica (%)", "Casa mais critica",
              "Utilizacao do trecho mais critico (%)", "Trecho mais critico",
              "Volume equivalente por dia (m3/dia)", "Vazao media requerida do sistema (m3/h)",
              "Capacidade instalada do sistema (m3/h)", "Utilizacao do sistema (%)",
              "Deficit de vazao do sistema (m3/h)", "Deficit de volume do sistema (m3)",
              "Deficit de horas do sistema (h)", "Chave de ordenacao"]
    pos["TOTAIS"] = {"titulo": linha, "cab": linha + 1, "ini": linha + 2,
                     "fim": linha + 1 + len(totais)}

    def col_sem(k):
        return gcl(3 + k)

    def escrever_cab(bloco):
        p = pos[bloco]
        ws.cell(row=p["titulo"], column=2, value=dict((b[0], b[1]) for b in blocos).get(
            bloco, "Indicadores do sistema RDM")).style = "secao"
        ws.cell(row=p["cab"], column=2, value="Semana ->").style = "cabecalho"
        for k in range(nsem):
            ws.cell(row=p["cab"], column=3 + k, value=k + 1).style = "cabecalho"
            ws.column_dimensions[col_sem(k)].width = 11
    for b in blocos:
        escrever_cab(b[0])
    escrever_cab("TOTAIS")
    ws.column_dimensions["B"].width = 40

    for chave, _, nomes, fmt in blocos:
        p = pos[chave]
        for i, nome in enumerate(nomes):
            ws.cell(row=p["ini"] + i, column=2, value=nome).style = "rotulo"

    vc, ac, hc, qc, uc, dc = (pos[k]["ini"] for k in
                              ("VOL_CASA", "AREA_CASA", "HORAS_CASA", "QREQ_CASA", "UTIL_CASA", "DEF_CASA"))
    vt, qt, qd, ut, dt_ = (pos[k]["ini"] for k in
                           ("VOL_TRECHO", "QREQ_TRECHO", "QDISP_TRECHO", "UTIL_TRECHO", "DEF_TRECHO"))
    tot = pos["TOTAIS"]["ini"]
    cab_tot = pos["TOTAIS"]["cab"]

    for k in range(nsem):
        C = col_sem(k)
        s_ref = "%s$%d" % (C, pos["VOL_CASA"]["cab"])
        horas_sem = "INDEX(%s,MATCH(%s,%s,0))" % (sem("K"), s_ref, sem("B"))
        for i in range(n):
            casa = "$B%d" % (vc + i)
            ws.cell(row=vc + i, column=3 + k,
                    value="=SUMIFS(%s,%s,%s,%s,%s)" % (ps("Z"), ps("F"), casa, ps("B"), s_ref)).style = "num0"
            ws.cell(row=ac + i, column=3 + k,
                    value='=SUMIFS(%s,%s,%s,%s,%s,%s,"<>SEM CULTURA")'
                          % (ps("H"), ps("F"), "$B%d" % (ac + i), ps("B"), s_ref, ps("K"))).style = "num1"
            ws.cell(row=hc + i, column=3 + k,
                    value="=SUMIFS(%s,%s,%s,%s,%s)" % (ps("AB"), ps("F"), "$B%d" % (hc + i), ps("B"), s_ref)
                    ).style = "num1"
            ws.cell(row=qc + i, column=3 + k,
                    value="=IFERROR(%s%d/%s,0)" % (C, vc + i, horas_sem)).style = "num1"
            cap = "INDEX(%s,MATCH($B%d,%s,0))" % (cas("H"), uc + i, cas("B"))
            ws.cell(row=uc + i, column=3 + k,
                    value="=IFERROR(%s%d/%s,0)" % (C, qc + i, cap)).style = "pct"
            ws.cell(row=dc + i, column=3 + k,
                    value="=IFERROR(%s%d-%s,0)" % (C, qc + i, cap)).style = "num1"
        for i in range(len(trechos)):
            linha_mapa = RT["mapa"] + i
            termos = "+".join("CADASTRO_TRECHOS_CANAL!$%s$%d*%s%d" % (gcl(3 + j), linha_mapa, C, vc + j)
                              for j in range(n))
            ws.cell(row=vt + i, column=3 + k, value="=" + termos).style = "num0"
            ws.cell(row=qt + i, column=3 + k,
                    value="=IFERROR(%s%d/%s,0)" % (C, vt + i, horas_sem)).style = "num1"
            ws.cell(row=qd + i, column=3 + k,
                    value="=CADASTRO_TRECHOS_CANAL!$K$%d" % (RT["ini"] + i)).style = "num1"
            ws.cell(row=ut + i, column=3 + k,
                    value="=IFERROR(%s%d/%s%d,0)" % (C, qt + i, C, qd + i)).style = "pct"
            ws.cell(row=dt_ + i, column=3 + k,
                    value="=%s%d-%s%d" % (C, qt + i, C, qd + i)).style = "num1"

        st = "%s$%d" % (C, cab_tot)
        hs = "INDEX(%s,MATCH(%s,%s,0))" % (sem("K"), st, sem("B"))
        area_ativa = 'SUMIFS(%s,%s,%s,%s,"<>SEM CULTURA")' % (ps("H"), ps("B"), st, ps("K"))
        f = [
            "=SUM(%s%d:%s%d)" % (C, vc, C, vc + n - 1),
            "=SUM(%s%d:%s%d)" % (C, ac, C, ac + n - 1),
            "=IFERROR(INDEX(%s,MATCH(%s,%s,0)),0)" % (sem("H"), st, sem("B")),
            "=IFERROR(SUMIFS(%s,%s,%s)/%s,0)" % (ps("AG"), ps("B"), st, area_ativa),
            "=IFERROR(SUMIFS(%s,%s,%s)/%s,0)" % (ps("AH"), ps("B"), st, area_ativa),
            "=IFERROR(SUMIFS(%s,%s,%s)/%s,0)" % (ps("AI"), ps("B"), st, area_ativa),
            "=IFERROR(%s%d/(%s%d*MM_PARA_M3_HA),0)" % (C, tot, C, tot + 1),
            "=SUM(%s%d:%s%d)" % (C, hc, C, hc + n - 1),
            "=IFERROR(_xlfn.MAXIFS(%s,%s,%s),0)" % (ps("AD"), ps("B"), st),
            '=IFERROR(INDEX(%s,MATCH(%s&"|"&TEXT(%s%d,"0.000000"),%s,0)),"-")'
            % (ps("E"), st, C, tot + 8, ps("AF")),
            "=MAX(%s%d:%s%d)" % (C, uc, C, uc + n - 1),
            "=IFERROR(INDEX($B$%d:$B$%d,MATCH(%s%d,%s%d:%s%d,0)),\"-\")"
            % (uc, uc + n - 1, C, tot + 10, C, uc, C, uc + n - 1),
            "=MAX(%s%d:%s%d)" % (C, ut, C, ut + len(trechos) - 1),
            "=IFERROR(INDEX($B$%d:$B$%d,MATCH(%s%d,%s%d:%s%d,0)),\"-\")"
            % (ut, ut + len(trechos) - 1, C, tot + 12, C, ut, C, ut + len(trechos) - 1),
            "=%s%d/DIAS_POR_SEMANA" % (C, tot),
            "=IFERROR(%s%d/%s,0)" % (C, tot, hs),
            "=SUM(%s)" % cas("H"),
            "=IFERROR(%s%d/%s%d,0)" % (C, tot + 15, C, tot + 16),
            "=MAX(0,%s%d-%s%d)" % (C, tot + 15, C, tot + 16),
            "=MAX(0,%s%d-%s%d*%s)" % (C, tot, C, tot + 16, hs),
            "=IFERROR(MAX(0,%s%d/%s%d-%s),0)" % (C, tot, C, tot + 16, hs),
            "=%s%d+%s/1000000" % (C, tot, st),
        ]
        fmts = ["num0", "num1", "num1", "num1", "num1", "num1", "num1", "num1", "pct", "texto",
                "pct", "texto", "pct", "texto", "num0", "num1", "num1", "pct", "num1", "num0",
                "num1", "num2"]
        for i, (v, fm) in enumerate(zip(f, fmts)):
            ws.cell(row=tot + i, column=3 + k, value=v).style = fm
    # ---- coluna de resumo da safra (MAX ou SOMA de cada linha)
    cr = 3 + nsem
    ws.cell(row=5, column=cr, value="Resumo da safra").style = "cabecalho"
    ws.column_dimensions[gcl(cr)].width = 17
    C0, C1 = gcl(3), gcl(2 + nsem)
    agreg = {"VOL_CASA": "SUM", "AREA_CASA": "MAX", "HORAS_CASA": "SUM", "QREQ_CASA": "MAX",
             "UTIL_CASA": "MAX", "DEF_CASA": "MAX", "VOL_TRECHO": "SUM", "QREQ_TRECHO": "MAX",
             "QDISP_TRECHO": "MAX", "UTIL_TRECHO": "MAX", "DEF_TRECHO": "MAX"}
    for chave, _, nomes, fmt in blocos:
        p_ = pos[chave]
        ws.cell(row=p_["cab"], column=cr, value=agreg[chave] + " da safra").style = "cabecalho"
        for i in range(len(nomes)):
            ws.cell(row=p_["ini"] + i, column=cr,
                    value="=%s(%s%d:%s%d)" % (agreg[chave], C0, p_["ini"] + i, C1, p_["ini"] + i)
                    ).style = fmt
    agreg_tot = ["SUM", "MAX", "SUM", "SUM", "SUM", "SUM", "SUM", "SUM", "MAX", "-", "MAX", "-",
                 "MAX", "-", "MAX", "MAX", "MAX", "MAX", "MAX", "SUM", "MAX", "MAX"]
    ws.cell(row=cab_tot, column=cr, value="Resumo da safra").style = "cabecalho"
    for i, (t, a) in enumerate(zip(totais, agreg_tot)):
        if a == "-":
            continue
        ws.cell(row=tot + i, column=cr,
                value="=%s(%s%d:%s%d)" % (a, C0, tot + i, C1, tot + i)).style = fmts[i]
    pos["col_resumo"] = cr

    for i, t in enumerate(totais):
        ws.cell(row=tot + i, column=2, value=t).style = "rotulo"
    ws.freeze_panes = "C6"
    pos["nsem"] = nsem
    pos["casas"] = casas
    pos["trechos"] = trechos
    return pos


def _aux(pos, bloco, i, col):
    return "AUX_MATRIZES!%s%d" % (col, pos[bloco]["ini"] + i)


# =========================================================== RESUMO_CASAS
def aba_resumo_casas(wb, D, AX):
    ws = wb.create_sheet("RESUMO_CASAS")
    est.cabecalho_pagina(ws, "RESUMO POR CASA DE BOMBA  -  matriz CASA x SEMANA",
                         "Escolha o indicador na celula amarela: mm, m3, m3/h, horas ou utilizacao (%). "
                         "A escala de cor acompanha o indicador escolhido.", "L")
    ind = ["m3 (volume da semana)", "mm (lamina bruta media)", "m3/h (vazao media requerida)",
           "horas (horas necessarias)", "% (utilizacao da capacidade)"]
    ws["B5"] = "Indicador exibido:"
    ws["B5"].style = "rotulo"
    ws["D5"] = ind[0]
    ws["D5"].style = "entrada"
    ws.merge_cells("D5:H5")
    for i, t in enumerate(ind):                       # lista de apoio (oculta)
        ws.cell(row=100 + i, column=40, value=t).style = "texto"
    ws.column_dimensions["AN"].hidden = True
    dv = DataValidation(type="list", formula1="=$AN$100:$AN$104", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("D5")
    lista = "$AN$100:$AN$104"

    nsem, casas = AX["nsem"], AX["casas"]
    cab = 7
    ws.cell(row=cab, column=2, value="Casa de bomba").style = "cabecalho"
    ws.column_dimensions["B"].width = 16
    for k in range(nsem):
        c = ws.cell(row=cab, column=3 + k, value="=SEMANAS!$B$%d" % (6 + k))
        c.style = "cabecalho"
        ws.column_dimensions[gcl(3 + k)].width = 9
    ws.cell(row=cab - 1, column=3,
            value='="Semanas da safra  ("&TEXT(SEMANAS!$C$6,"dd/mm/yy")&" a "&TEXT(SEMANAS!$D$%d,"dd/mm/yy")&")"'
                  % (5 + nsem)).style = "secao"
    for i, casa in enumerate(casas):
        r = cab + 1 + i
        ws.cell(row=r, column=2, value=casa).style = "rotulo"
        for k in range(nsem):
            C = gcl(3 + k)
            ws.cell(row=r, column=3 + k,
                    value="=IFERROR(CHOOSE(MATCH($D$5,%s,0),%s,IFERROR(%s/(%s*MM_PARA_M3_HA),0),%s,%s,%s*100),0)"
                          % (lista, _aux(AX, "VOL_CASA", i, C), _aux(AX, "VOL_CASA", i, C),
                             _aux(AX, "AREA_CASA", i, C), _aux(AX, "QREQ_CASA", i, C),
                             _aux(AX, "HORAS_CASA", i, C), _aux(AX, "UTIL_CASA", i, C))
                    ).style = "num1"
    fim_l, fim_c = cab + len(casas), gcl(2 + nsem)
    faixa = "C%d:%s%d" % (cab + 1, fim_c, fim_l)
    ws.conditional_formatting.add(faixa, ColorScaleRule(
        start_type="min", start_color="FFFFFF", mid_type="percentile", mid_value=50,
        mid_color="FFEB9C", end_type="max", end_color="F8696B"))
    ws.freeze_panes = "C%d" % (cab + 1)
    nota(ws, fim_l + 2,
         "Leitura: cada celula e a demanda da casa naquela semana, somando todos os pivos ligados a ela "
         "(vinculo definido em CADASTRO_PIVOS). Para o mapa de calor de utilizacao com faixas fixas, "
         "veja a aba MAPA_CALOR.", 12)
    return {"cab": cab, "ini": cab + 1, "fim": fim_l}


# ========================================================= RESUMO_TRECHOS
def aba_resumo_trechos(wb, D, AX):
    ws = wb.create_sheet("RESUMO_TRECHOS")
    est.cabecalho_pagina(ws, "RESUMO POR TRECHO DE CANAL  -  matriz TRECHO x SEMANA",
                         "Demanda acumulada a jusante de cada trecho, conforme a matriz TRECHO x CASA do "
                         "CADASTRO_TRECHOS_CANAL. Deficit positivo = trecho hidraulicamente limitante.", "L")
    nsem, trechos = AX["nsem"], AX["trechos"]
    blocos = [("Demanda do trecho (m3/semana)", "VOL_TRECHO", "num0"),
              ("Vazao requerida (m3/h)", "QREQ_TRECHO", "num1"),
              ("Capacidade disponivel (m3/h)", "QDISP_TRECHO", "num1"),
              ("Utilizacao (%)", "UTIL_TRECHO", "pct"),
              ("Deficit de vazao (m3/h)  -  positivo = TRECHO LIMITANTE", "DEF_TRECHO", "num1")]
    linha = 5
    ws.column_dimensions["B"].width = 16
    for titulo, chave, fmt in blocos:
        ws.cell(row=linha, column=2, value=titulo).style = "secao"
        ws.cell(row=linha + 1, column=2, value="Trecho").style = "cabecalho"
        for k in range(nsem):
            ws.cell(row=linha + 1, column=3 + k, value="=SEMANAS!$B$%d" % (6 + k)).style = "cabecalho"
            ws.column_dimensions[gcl(3 + k)].width = 9
        for i, t in enumerate(trechos):
            r = linha + 2 + i
            ws.cell(row=r, column=2, value=t).style = "rotulo"
            for k in range(nsem):
                ws.cell(row=r, column=3 + k,
                        value="=%s" % _aux(AX, chave, i, gcl(3 + k))).style = fmt
        faixa = "C%d:%s%d" % (linha + 2, gcl(2 + nsem), linha + 1 + len(trechos))
        if chave == "UTIL_TRECHO":
            ws.conditional_formatting.add(faixa, ColorScaleRule(
                start_type="num", start_value=0, start_color="C6EFCE",
                mid_type="num", mid_value=0.9, mid_color="FFEB9C",
                end_type="num", end_value=1, end_color="F8696B"))
        elif chave == "DEF_TRECHO":
            ws.conditional_formatting.add(faixa, CellIsRule(
                operator="greaterThan", formula=["0"],
                fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))
        linha += len(trechos) + 4
    nota(ws, linha, "Enquanto a capacidade operacional de cada trecho nao for cadastrada, a linha 'Capacidade "
                    "disponivel' repete a soma das casas atendidas: a utilizacao fica em 100% e o deficit em "
                    "zero por construcao. Preencha CADASTRO_TRECHOS_CANAL para que a analise tenha sentido.", 12)


# ============================================================== MAPA_CALOR
def aba_mapa_calor(wb, D, AX):
    ws = wb.create_sheet("MAPA_CALOR")
    est.cabecalho_pagina(ws, "MAPA DE CALOR  -  utilizacao da capacidade hidraulica",
                         "Verde = folga | Amarelo = atencao | Laranja = proximo do limite | "
                         "Vermelho = capacidade excedida. As faixas vem de PARAMETROS_GERAIS.", "L")
    nsem = AX["nsem"]
    linha = 5
    ws.column_dimensions["B"].width = 16
    for titulo, chave, nomes in (("CASA DE BOMBA x SEMANA", "UTIL_CASA", AX["casas"]),
                                 ("TRECHO DE CANAL x SEMANA", "UTIL_TRECHO", AX["trechos"])):
        ws.cell(row=linha, column=2, value=titulo).style = "secao"
        ws.cell(row=linha + 1, column=2, value="Semana ->").style = "cabecalho"
        for k in range(nsem):
            ws.cell(row=linha + 1, column=3 + k, value="=SEMANAS!$B$%d" % (6 + k)).style = "cabecalho"
            ws.column_dimensions[gcl(3 + k)].width = 6
        for i, nome in enumerate(nomes):
            r = linha + 2 + i
            ws.cell(row=r, column=2, value=nome).style = "rotulo"
            for k in range(nsem):
                ws.cell(row=r, column=3 + k, value="=%s" % _aux(AX, chave, i, gcl(3 + k))).style = "pct"
        faixa = "C%d:%s%d" % (linha + 2, gcl(2 + nsem), linha + 1 + len(nomes))
        for op, ref, cor in (("greaterThan", "LIMITE_CRITICO", "F8696B"),
                             ("greaterThan", "LIMITE_ATENCAO", "F4B183"),
                             ("greaterThan", "LIMITE_CONFORTAVEL", "FFEB9C"),
                             ("lessThanOrEqual", "LIMITE_CONFORTAVEL", "C6EFCE")):
            ws.conditional_formatting.add(faixa, CellIsRule(
                operator=op, formula=[ref], fill=est.PatternFill("solid", fgColor=cor)))
        linha += len(nomes) + 4
    nota(ws, linha, "As quatro regras sao avaliadas em ordem: acima de LIMITE_CRITICO (100%) vermelho, "
                    "acima de LIMITE_ATENCAO (90%) laranja, acima de LIMITE_CONFORTAVEL (80%) amarelo, "
                    "o resto verde. Mude os limites em PARAMETROS_GERAIS e o mapa inteiro se reajusta.", 12)


# =========================================================== CAPACIDADE_21H
def aba_capacidade(wb, D, AX, RS):
    ws = wb.create_sheet("CAPACIDADE_21H")
    est.cabecalho_pagina(ws, "CAPACIDADE_21H  -  a infraestrutura cabe em 21 horas por dia?",
                         "Uma linha por semana. Confronta a demanda com a capacidade disponivel em "
                         "HORAS_OPERACAO_DIA x DIAS_POR_SEMANA horas, nos quatro niveis: pivo, casa de bomba, "
                         "trecho de canal e sistema RDM.", "X")
    cols = ["Semana", "Data inicial", "Data final", "Volume necessario (m3)",
            "Capacidade maxima da semana (m3)", "Utilizacao do sistema (%)", "Deficit de volume (m3)",
            "Deficit de vazao (m3/h)", "Deficit de horas (h)", "Horas adicionais necessarias (h)",
            "Vazao adicional necessaria (m3/h)", "Status do sistema",
            "Utilizacao max. de pivo (%)", "Pivo mais critico", "Status do pivo",
            "Utilizacao da casa mais critica (%)", "Casa mais critica", "Status da casa",
            "Utilizacao do trecho mais critico (%)", "Trecho mais critico", "Status do trecho",
            "Chave de ordenacao"]
    escrever_tabela(ws, cols, [8, 12, 12, 15, 16, 14, 15, 14, 13, 15, 16, 15, 14, 13, 14,
                               16, 14, 14, 17, 14, 14, 15])
    tot = AX["TOTAIS"]["ini"]
    nsem = AX["nsem"]

    def T(off, k):
        return "AUX_MATRIZES!%s%d" % (gcl(3 + k), tot + off)

    st = lambda ref: ('=IF(%s>LIMITE_CRITICO,"INVIAVEL - demanda excede a capacidade em 21 h/dia",'   # noqa: E731
                      'IF(%s>LIMITE_ATENCAO,"CRITICO - utilizacao acima de "&TEXT(LIMITE_ATENCAO,"0%%"),'
                      'IF(%s>LIMITE_CONFORTAVEL,"ATENCAO - utilizacao acima de "&TEXT(LIMITE_CONFORTAVEL,"0%%"),'
                      '"OK - capacidade suficiente")))' % (ref, ref, ref))
    for k in range(nsem):
        r = DAT + k
        vals = [
            "=SEMANAS!$B$%d" % (RS["ini"] + k),
            "=SEMANAS!$C$%d" % (RS["ini"] + k),
            "=SEMANAS!$D$%d" % (RS["ini"] + k),
            "=%s" % T(0, k),
            "=%s*SEMANAS!$K$%d" % (T(16, k), RS["ini"] + k),
            "=IFERROR($E%d/$F%d,0)" % (r, r),
            "=MAX(0,$E%d-$F%d)" % (r, r),
            "=%s" % T(18, k),
            "=%s" % T(20, k),
            "=IFERROR(MAX(0,$E%d/%s-SEMANAS!$K$%d),0)" % (r, T(16, k), RS["ini"] + k),
            "=IFERROR(MAX(0,$E%d/SEMANAS!$K$%d-%s),0)" % (r, RS["ini"] + k, T(16, k)),
            st("$G%d" % r),
            "=%s" % T(8, k),
            "=%s" % T(9, k),
            st("$N%d" % r),
            "=%s" % T(10, k),
            "=%s" % T(11, k),
            st("$Q%d" % r),
            "=%s" % T(12, k),
            "=%s" % T(13, k),
            st("$T%d" % r),
            "=%s" % T(21, k),
        ]
        fmts = ["num0", "data", "data", "num0", "num0", "pct", "num0", "num1", "num1", "num1",
                "num1", "texto", "pct", "texto", "texto", "pct", "texto", "texto", "pct",
                "texto", "texto", "num2"]
        for j, (v, f) in enumerate(zip(vals, fmts)):
            ws.cell(row=r, column=2 + j, value=v).style = f
    fim = DAT + nsem - 1
    for col in ("M", "P", "S", "V"):
        for txt, cor in (("OK", est.VERDE_OK), ("ATENCAO", "FFEB9C"),
                         ("CRITICO", est.LARANJA), ("INVIAVEL", est.VERMELHO_PEND)):
            ws.conditional_formatting.add("%s%d:%s%d" % (col, DAT, col, fim), CellIsRule(
                operator="beginsWith", formula=['"%s"' % txt],
                fill=est.PatternFill("solid", fgColor=cor)))
    ws.column_dimensions["W"].hidden = True
    linha = fim + 2
    ws.cell(row=linha, column=2, value="TOTAIS DA SAFRA").style = "secao"
    for j, (rot, form, fmt) in enumerate((
            ("Volume necessario na safra (m3)", "=SUM($E$%d:$E$%d)" % (DAT, fim), "num0"),
            ("Semana de pico de volume", "=INDEX($B$%d:$B$%d,MATCH(MAX($E$%d:$E$%d),$E$%d:$E$%d,0))"
             % (DAT, fim, DAT, fim, DAT, fim), "num0"),
            ("Volume da semana de pico (m3)", "=MAX($E$%d:$E$%d)" % (DAT, fim), "num0"),
            ("Maior utilizacao do sistema (%)", "=MAX($G$%d:$G$%d)" % (DAT, fim), "pct"),
            ("Semanas com utilizacao acima de 100%",
             "=COUNTIF($G$%d:$G$%d,\">\"&LIMITE_CRITICO)" % (DAT, fim), "num0"),
            ("Deficit total de volume na safra (m3)", "=SUM($H$%d:$H$%d)" % (DAT, fim), "num0"))):
        ws.cell(row=linha + 1 + j, column=2, value=rot).style = "rotulo"
        ws.cell(row=linha + 1 + j, column=6, value=form).style = fmt
    return {"ini": DAT, "fim": fim}


# ============================================================ TOP10_SEMANAS
def aba_top10(wb, D, AX, RCP):
    ws = wb.create_sheet("TOP10_SEMANAS")
    est.cabecalho_pagina(ws, "TOP 10 SEMANAS CRITICAS DA SAFRA",
                         "Ranking automatico pelo volume de irrigacao necessario na semana. "
                         "Recalcula sozinho a cada mudanca de plantio, clima, Kc ou vazao.", "N")
    escrever_tabela(ws, ["Posicao", "Semana", "Data inicial", "ETo (mm)", "ETc (mm)", "Chuva util (mm)",
                         "Area irrigada ativa (ha)", "Volume necessario (m3)",
                         "Vazao media requerida (m3/h)", "Utilizacao do sistema (%)",
                         "Casa mais critica", "Trecho mais critico"],
                     [9, 9, 13, 11, 11, 13, 16, 16, 17, 16, 14, 15])
    ci, cf = RCP["ini"], RCP["fim"]
    chave = "CAPACIDADE_21H!$W$%d:$W$%d" % (ci, cf)
    semanas = "CAPACIDADE_21H!$B$%d:$B$%d" % (ci, cf)
    tot, cab = AX["TOTAIS"]["ini"], AX["TOTAIS"]["cab"]
    ult = gcl(2 + AX["nsem"])

    def aux_linha(off, r):
        return ("INDEX(AUX_MATRIZES!$C$%d:$%s$%d,MATCH($C%d,AUX_MATRIZES!$C$%d:$%s$%d,0))"
                % (tot + off, ult, tot + off, r, cab, ult, cab))

    for k in range(10):
        r = DAT + k
        mc = "MATCH($C%d,%s,0)" % (r, semanas)
        ws.cell(row=r, column=2, value=k + 1).style = "num0"
        ws.cell(row=r, column=3,
                value="=IFERROR(INDEX(%s,MATCH(LARGE(%s,$B%d),%s,0)),\"-\")" % (semanas, chave, r, chave)
                ).style = "num0"
        ws.cell(row=r, column=4,
                value="=IFERROR(INDEX(CAPACIDADE_21H!$C$%d:$C$%d,%s),\"\")" % (ci, cf, mc)).style = "data"
        for j, off in enumerate((2, 3, 4, 1, 0, 15)):
            ws.cell(row=r, column=5 + j, value="=IFERROR(%s,0)" % aux_linha(off, r)
                    ).style = ["num1", "num1", "num1", "num1", "num0", "num1"][j]
        ws.cell(row=r, column=11,
                value="=IFERROR(INDEX(CAPACIDADE_21H!$G$%d:$G$%d,%s),0)" % (ci, cf, mc)).style = "pct"
        ws.cell(row=r, column=12, value="=IFERROR(%s,\"-\")" % aux_linha(11, r)).style = "texto"
        ws.cell(row=r, column=13, value="=IFERROR(%s,\"-\")" % aux_linha(13, r)).style = "texto"
    nota(ws, DAT + 11,
         "A chave de ordenacao soma o volume da semana com semana/1.000.000 para desempatar semanas "
         "de volume identico sem alterar a ordem de grandeza.", 13)


# ======================================================== BALANCO_CULTURA
def aba_balanco_cultura(wb, D, RPS, RR):
    ws = wb.create_sheet("BALANCO_CULTURA")
    est.cabecalho_pagina(ws, "BALANCO DA SAFRA POR CULTURA",
                         "Area plantada acumulada = soma das areas dos ciclos da cultura na safra; um mesmo "
                         "pivo que recebe a cultura em dois ciclos entra duas vezes. As laminas em mm sao "
                         "ponderadas por essa mesma base, entao representam mm por hectare-ciclo.",
                         "P")
    escrever_tabela(ws, ["Cultura", "Area plantada acumulada (ha)", "Ciclos na safra", "ETo durante a ocupacao (mm)",
                         "Kc medio ponderado", "ETc acumulada (mm)", "Chuva util (mm)",
                         "Irrigacao liquida (mm)", "Irrigacao bruta (mm)", "Volume (m3)",
                         "mm irrigados por ha", "Horas de irrigacao (h)"],
                    [18, 15, 13, 18, 14, 15, 13, 15, 15, 15, 14, 16])
    culturas = sorted({c["cultura"] for c in D["ciclos"]})
    ps = lambda col: "PLANEJAMENTO_SEMANAL!$%s$%d:$%s$%d" % (col, RPS["ini"], col, RPS["fim"])  # noqa: E731
    rot = lambda col: "ROTACAO_RDM!$%s$%d:$%s$%d" % (col, RR["ini"], col, RR["fim"])            # noqa: E731
    for i, cult in enumerate(culturas):
        r = DAT + i
        ws.cell(row=r, column=2, value=cult).style = "rotulo"
        ws.cell(row=r, column=3, value="=SUMIF(%s,$B%d,%s)" % (rot("F"), r, rot("E"))).style = "num2"
        ws.cell(row=r, column=4, value="=COUNTIF(%s,$B%d)" % (rot("F"), r)).style = "num0"
        for j, col in enumerate(("AJ", "AG", "AH", "AI")):
            dest = {0: 5, 1: 7, 2: 8, 3: 9}[j]
            ws.cell(row=r, column=dest,
                    value="=IFERROR(SUMIF(%s,$B%d,%s)/$C%d,0)" % (ps("K"), r, ps(col), r)).style = "num1"
        ws.cell(row=r, column=6, value="=IFERROR($G%d/$E%d,0)" % (r, r)).style = "num3"
        ws.cell(row=r, column=11, value="=SUMIF(%s,$B%d,%s)" % (ps("K"), r, ps("Z"))).style = "num0"
        ws.cell(row=r, column=10, value="=IFERROR($K%d/($C%d*MM_PARA_M3_HA),0)" % (r, r)).style = "num1"
        ws.cell(row=r, column=12, value="=$J%d" % r).style = "num1"
        ws.cell(row=r, column=13, value="=SUMIF(%s,$B%d,%s)" % (ps("K"), r, ps("AB"))).style = "num1"
    fim = DAT + len(culturas) - 1
    ws.cell(row=fim + 1, column=2, value="TOTAL RDM").style = "rotulo"
    for col in (3, 4, 11, 13):
        ws.cell(row=fim + 1, column=col,
                value="=SUM(%s%d:%s%d)" % (gcl(col), DAT, gcl(col), fim)).style = "num0"
    nota(ws, fim + 3,
         "Soja e algodao sao as duas culturas que concentram a demanda da RDM: compare as linhas SOJA e "
         "ALGODAO em 'Volume (m3)' e em 'mm irrigados por ha'.", 13)


# =========================================================== BALANCO_PIVO
def aba_balanco_pivo(wb, D, RPS, RR, RP):
    ws = wb.create_sheet("BALANCO_PIVO")
    est.cabecalho_pagina(ws, "BALANCO DA SAFRA POR PIVO",
                         "Responde diretamente: quantas horas cada pivo precisa irrigar na safra inteira.", "S")
    escrever_tabela(ws, ["Pivo", "Casa de bomba", "Trecho de canal", "Area (ha)", "Cultura 1", "Cultura 2",
                         "Cultura 3", "Semanas com cultura", "ETc total (mm)", "Chuva util total (mm)",
                         "Irrigacao liquida (mm)", "Lamina bruta total (mm)", "Volume total (m3)",
                         "Vazao do pivo (m3/h)", "Horas totais de funcionamento (h)",
                         "Utilizacao maxima semanal (%)", "Semana de pico", "Status"],
                    [10, 13, 13, 10, 16, 16, 16, 13, 13, 15, 14, 15, 15, 13, 18, 15, 12, 26])
    ps = lambda col: "PLANEJAMENTO_SEMANAL!$%s$%d:$%s$%d" % (col, RPS["ini"], col, RPS["fim"])  # noqa: E731
    rot = lambda col: "ROTACAO_RDM!$%s$%d:$%s$%d" % (col, RR["ini"], col, RR["fim"])            # noqa: E731
    piv = lambda col: "CADASTRO_PIVOS!$%s$%d:$%s$%d" % (col, RP["ini"], col, RP["fim"])         # noqa: E731
    for i, p in enumerate(D["pivos"]):
        r = DAT + i
        mp = "MATCH($B%d,%s,0)" % (r, piv("B"))
        prim = "MATCH($B%d,%s,0)" % (r, rot("D"))
        ws.cell(row=r, column=2, value=p["pivo"]).style = "rotulo"
        ws.cell(row=r, column=3, value="=IFERROR(INDEX(%s,%s),\"?\")" % (piv("C"), mp)).style = "texto"
        ws.cell(row=r, column=4, value="=IFERROR(INDEX(%s,%s),\"?\")" % (piv("D"), mp)).style = "texto"
        ws.cell(row=r, column=5, value="=IFERROR(INDEX(%s,%s),0)" % (piv("E"), mp)).style = "num2"
        for j in range(3):
            ws.cell(row=r, column=6 + j,
                    value='=IFERROR(IF(INDEX(%s,%s+%d)=$B%d,INDEX(%s,%s+%d),""),"")'
                          % (rot("D"), prim, j, r, rot("F"), prim, j)).style = "texto"
        ws.cell(row=r, column=9,
                value='=COUNTIFS(%s,$B%d,%s,"<>SEM CULTURA")' % (ps("E"), r, ps("K"))).style = "num0"
        for j, col in enumerate(("AG", "AH", "AI")):
            ws.cell(row=r, column=10 + j,
                    value="=IFERROR(SUMIF(%s,$B%d,%s)/$E%d,0)" % (ps("E"), r, ps(col), r)).style = "num1"
        ws.cell(row=r, column=13, value="=SUMIF(%s,$B%d,%s)" % (ps("E"), r, ps("Y"))).style = "num1"
        ws.cell(row=r, column=14, value="=SUMIF(%s,$B%d,%s)" % (ps("E"), r, ps("Z"))).style = "num0"
        ws.cell(row=r, column=15, value="=IFERROR(INDEX(%s,%s),0)" % (piv("I"), mp)).style = "num1"
        ws.cell(row=r, column=16, value="=SUMIF(%s,$B%d,%s)" % (ps("E"), r, ps("AB"))).style = "num1"
        ws.cell(row=r, column=17, value="=IFERROR(_xlfn.MAXIFS(%s,%s,$B%d),0)" % (ps("AD"), ps("E"), r)
                ).style = "pct"
        ws.cell(row=r, column=18,
                value='=IFERROR(INDEX(%s,MATCH($B%d&"|"&TEXT($Q%d,"0.000000"),%s,0)),"-")'
                      % (ps("B"), r, r, ps("AK"))).style = "num0"
        ws.cell(row=r, column=19,
                value='=IF($Q%d>LIMITE_CRITICO,"INVIAVEL em 21 h/dia",IF($Q%d>LIMITE_ATENCAO,"CRITICO",'
                      'IF($Q%d>LIMITE_CONFORTAVEL,"ATENCAO","OK")))' % (r, r, r)).style = "texto"
    fim = DAT + len(D["pivos"]) - 1
    for txt, cor in (("OK", est.VERDE_OK), ("ATENCAO", "FFEB9C"),
                     ("CRITICO", est.LARANJA), ("INVIAVEL", est.VERMELHO_PEND)):
        ws.conditional_formatting.add("S%d:S%d" % (DAT, fim), CellIsRule(
            operator="beginsWith", formula=['"%s"' % txt],
            fill=est.PatternFill("solid", fgColor=cor)))
    ws.cell(row=fim + 1, column=2, value="TOTAL RDM").style = "rotulo"
    for col in (5, 14, 16):
        ws.cell(row=fim + 1, column=col,
                value="=SUM(%s%d:%s%d)" % (gcl(col), DAT, gcl(col), fim)).style = "num1"


# ========================================================== DASHBOARD_RDM
def aba_dashboard(wb, D, AX, RPS, RP, RCP, BC):
    ws = wb.create_sheet("DASHBOARD_RDM")
    est.cabecalho_pagina(ws, "DASHBOARD RDM  -  planejamento hidrico da safra",
                         "Todos os numeros abaixo sao formulas. Mude plantio, ETo, chuva, Kc, vazao ou horas "
                         "disponiveis e o painel inteiro se refaz.", "R")
    ps = lambda col: "PLANEJAMENTO_SEMANAL!$%s$%d:$%s$%d" % (col, RPS["ini"], col, RPS["fim"])  # noqa: E731
    piv = lambda col: "CADASTRO_PIVOS!$%s$%d:$%s$%d" % (col, RP["ini"], col, RP["fim"])         # noqa: E731
    cr = gcl(AX["col_resumo"])
    tot = AX["TOTAIS"]["ini"]
    ult = gcl(2 + AX["nsem"])
    cab = AX["TOTAIS"]["cab"]
    ci, cf = RCP["ini"], RCP["fim"]
    area_total = "SUM(%s)" % piv("E")

    def linha_aux(off, agreg="MAX"):
        return "%s(AUX_MATRIZES!$C$%d:$%s$%d)" % (agreg, tot + off, ult, tot + off)

    kpis = [
        ("Area irrigada cadastrada", "=%s" % area_total, "kpiValor1", "ha"),
        ("Area irrigada ativa no pico", "=%s" % linha_aux(1), "kpiValor1", "ha"),
        ("Demanda ETc da safra", "=IFERROR(SUM(%s)/%s,0)" % (ps("AG"), area_total), "kpiValor", "mm"),
        ("Demanda ETc da safra", "=SUM(%s)*MM_PARA_M3_HA" % ps("AG"), "kpiValor", "m3"),
        ("Chuva util da safra", "=IFERROR(SUM(%s)/%s,0)" % (ps("AH"), area_total), "kpiValor", "mm"),
        ("Chuva util da safra", "=SUM(%s)*MM_PARA_M3_HA" % ps("AH"), "kpiValor", "m3"),
        ("Irrigacao liquida necessaria", "=IFERROR(SUM(%s)/%s,0)" % (ps("AI"), area_total), "kpiValor", "mm"),
        ("Irrigacao liquida necessaria", "=SUM(%s)*MM_PARA_M3_HA" % ps("AI"), "kpiValor", "m3"),
        ("Irrigacao bruta necessaria", "=IFERROR(SUM(%s)/(%s*MM_PARA_M3_HA),0)" % (ps("Z"), area_total),
         "kpiValor", "mm"),
        ("Irrigacao bruta necessaria (VOLUME TOTAL DA SAFRA)", "=SUM(%s)" % ps("Z"), "kpiValor", "m3"),
        ("Pico de demanda", "=%s" % linha_aux(14), "kpiValor", "m3/dia"),
        ("Pico de vazao requerida", "=%s" % linha_aux(15), "kpiValor", "m3/h"),
        ("Capacidade instalada do sistema", "=%s" % linha_aux(16), "kpiValor", "m3/h"),
        ("Semana critica",
         "=IFERROR(INDEX(CAPACIDADE_21H!$B$%d:$B$%d,MATCH(MAX(CAPACIDADE_21H!$E$%d:$E$%d),"
         "CAPACIDADE_21H!$E$%d:$E$%d,0)),\"-\")" % (ci, cf, ci, cf, ci, cf), "kpiValorTxt", "semana"),
        ("Data da semana critica",
         "=IFERROR(INDEX(CAPACIDADE_21H!$C$%d:$C$%d,MATCH(MAX(CAPACIDADE_21H!$E$%d:$E$%d),"
         "CAPACIDADE_21H!$E$%d:$E$%d,0)),\"-\")" % (ci, cf, ci, cf, ci, cf), "kpiValorData", "data"),
        ("Casa de bomba mais critica",
         "=IFERROR(INDEX(AUX_MATRIZES!$B$%d:$B$%d,MATCH(MAX(AUX_MATRIZES!$%s$%d:$%s$%d),"
         "AUX_MATRIZES!$%s$%d:$%s$%d,0)),\"-\")"
         % (AX["UTIL_CASA"]["ini"], AX["UTIL_CASA"]["fim"], cr, AX["UTIL_CASA"]["ini"], cr,
            AX["UTIL_CASA"]["fim"], cr, AX["UTIL_CASA"]["ini"], cr, AX["UTIL_CASA"]["fim"]),
         "kpiValorTxt", "casa"),
        ("Trecho de canal mais critico",
         "=IFERROR(INDEX(AUX_MATRIZES!$B$%d:$B$%d,MATCH(MAX(AUX_MATRIZES!$%s$%d:$%s$%d),"
         "AUX_MATRIZES!$%s$%d:$%s$%d,0)),\"-\")"
         % (AX["UTIL_TRECHO"]["ini"], AX["UTIL_TRECHO"]["fim"], cr, AX["UTIL_TRECHO"]["ini"], cr,
            AX["UTIL_TRECHO"]["fim"], cr, AX["UTIL_TRECHO"]["ini"], cr, AX["UTIL_TRECHO"]["fim"]),
         "kpiValorTxt", "trecho"),
        ("Maior utilizacao hidraulica (sistema)", "=%s" % linha_aux(17), "kpiValorPct", "%"),
        ("Maior utilizacao de uma casa de bomba",
         "=MAX(AUX_MATRIZES!$%s$%d:$%s$%d)" % (cr, AX["UTIL_CASA"]["ini"], cr, AX["UTIL_CASA"]["fim"]),
         "kpiValorPct", "%"),
        ("Maior utilizacao de um trecho de canal",
         "=MAX(AUX_MATRIZES!$%s$%d:$%s$%d)" % (cr, AX["UTIL_TRECHO"]["ini"], cr, AX["UTIL_TRECHO"]["fim"]),
         "kpiValorPct", "%"),
        ("Maior utilizacao de um pivo", "=%s" % linha_aux(8), "kpiValorPct", "%"),
        ("Semanas inviaveis em 21 h/dia",
         "=COUNTIF(CAPACIDADE_21H!$G$%d:$G$%d,\">\"&LIMITE_CRITICO)" % (ci, cf), "kpiValor", "semanas"),
        ("Horas totais de irrigacao na safra", "=SUM(%s)" % ps("AB"), "kpiValor", "h"),
        ("Deficit total de volume na safra", "=SUM(CAPACIDADE_21H!$H$%d:$H$%d)" % (ci, cf), "kpiValor", "m3"),
        ("Casas de bomba com captacao insuficiente",
         '=COUNTIF(BALANCO_CASAS!$K$%d:$K$%d,"CAPTACAO*")' % (BC["resumo_ini"], BC["resumo_fim"]),
         "kpiValor", "de 7"),
        ("Consumo total das casas na safra",
         "=SUM(BALANCO_CASAS!$D$%d:$D$%d)" % (BC["resumo_ini"], BC["resumo_fim"]), "kpiValor", "m3"),
    ]
    linha = 5
    for i, (rot, form, stl, un) in enumerate(kpis):
        col = 2 + (i % 4) * 4
        if i % 4 == 0 and i:
            linha += 4
        ws.cell(row=linha, column=col, value=rot).style = "kpiRotulo"
        ws.merge_cells(start_row=linha, start_column=col, end_row=linha, end_column=col + 2)
        ws.row_dimensions[linha].height = 30
        ws.cell(row=linha + 1, column=col, value=form).style = stl
        ws.merge_cells(start_row=linha + 1, start_column=col, end_row=linha + 1, end_column=col + 1)
        ws.row_dimensions[linha + 1].height = 24
        ws.cell(row=linha + 1, column=col + 2, value=un).style = "texto"
    for c in range(2, 18):
        ws.column_dimensions[gcl(c)].width = 13

    linha += 4
    ws.cell(row=linha, column=2, value="RESPOSTA DO MODELO").style = "secao"
    ws.cell(row=linha + 1, column=2,
            value='="GESTAO DE AGUA - "&IF(PAINEL_GESTAO!$F$15=0,'
                  '"a captacao atende o consumo nas 7 casas o ano todo.",'
                  '"a captacao NAO acompanha o consumo em "&PAINEL_GESTAO!$F$15&'
                  '" casa(s). A mais critica e "&PAINEL_GESTAO!$F$18&", com "&'
                  'TEXT(PAINEL_GESTAO!$F$17,"#,##0")&" m3 faltantes na safra. Detalhe em PAINEL_GESTAO.")&'
                  '"   |   PIVOS EM 21 h/dia - cabem com folga: utilizacao maxima de "&'
                  'TEXT(MAX(PLANEJAMENTO_SEMANAL!$AD$%d:$AD$%d),"0.0%%")&'
                  '". O gargalo da RDM esta na captacao/reservatorio, nao nas horas de pivo."'
                  % (RPS["ini"], RPS["fim"])).style = "aviso"
    ws.merge_cells(start_row=linha + 1, start_column=2, end_row=linha + 2, end_column=17)
    ws.row_dimensions[linha + 1].height = 30
    nota(ws, linha + 4,
         "ATENCAO: enquanto ETo, chuva util, eficiencia, vazao operacional das casas e capacidade dos trechos "
         "estiverem como DADO PENDENTE (ver aba PENDENCIAS_CADASTRO), os numeros acima sao ordem de grandeza, "
         "nao compromisso operacional.", 17)
    return linha


# ================================================================ GRAFICOS
def aba_graficos(wb, D, AX, RS):
    ws = wb.create_sheet("GRAFICOS")
    est.cabecalho_pagina(ws, "GRAFICOS DA SAFRA RDM",
                         "Series alimentadas por formulas. Bloco de dados no rodape da aba (linhas 60 em diante).",
                         "N")
    nsem = AX["nsem"]
    tot = AX["TOTAIS"]["ini"]
    base = 60
    ws.cell(row=base - 1, column=2, value="BLOCO DE DADOS DOS GRAFICOS").style = "secao"
    ws.cell(row=base, column=2, value="Semana").style = "cabecalho"
    for k in range(nsem):
        ws.cell(row=base, column=3 + k, value="=SEMANAS!$B$%d" % (RS["ini"] + k)).style = "cabecalho"

    def serie(linha, rotulo, gerador, fmt="num1"):
        ws.cell(row=linha, column=2, value=rotulo).style = "rotulo"
        for k in range(nsem):
            ws.cell(row=linha, column=3 + k, value=gerador(k)).style = fmt

    A = lambda off, k: "=AUX_MATRIZES!%s%d" % (gcl(3 + k), tot + off)   # noqa: E731
    serie(base + 1, "ETo (mm/semana)", lambda k: A(2, k))
    serie(base + 2, "ETc (mm/semana)", lambda k: A(3, k))
    serie(base + 3, "Chuva util (mm/semana)", lambda k: A(4, k))
    serie(base + 4, "Necessidade de irrigacao liquida (mm/semana)", lambda k: A(5, k))
    serie(base + 5, "Lamina bruta de irrigacao (mm/semana)", lambda k: A(6, k))
    serie(base + 6, "Volume total demandado pela cultura (m3)",
          lambda k: "=(AUX_MATRIZES!%s%d+AUX_MATRIZES!%s%d)*AUX_MATRIZES!%s%d*MM_PARA_M3_HA"
                    % (gcl(3 + k), tot + 4, gcl(3 + k), tot + 5, gcl(3 + k), tot + 1), "num0")
    serie(base + 7, "Volume suprido pela chuva util (m3)",
          lambda k: "=AUX_MATRIZES!%s%d*AUX_MATRIZES!%s%d*MM_PARA_M3_HA"
                    % (gcl(3 + k), tot + 4, gcl(3 + k), tot + 1), "num0")
    serie(base + 8, "Volume necessario via irrigacao (m3)", lambda k: A(0, k), "num0")
    lin_casa = base + 10
    ws.cell(row=lin_casa - 1, column=2, value="Volume semanal por casa de bomba (m3)").style = "secao"
    for i, casa in enumerate(AX["casas"]):
        serie(lin_casa + i, casa,
              lambda k, i=i: "=%s" % _aux(AX, "VOL_CASA", i, gcl(3 + k)), "num0")
    lin_tr = lin_casa + len(AX["casas"]) + 2
    ws.cell(row=lin_tr - 1, column=2, value="Trechos de canal: vazao requerida x disponivel (m3/h)").style = "secao"
    linhas_tr = []
    for i, t in enumerate(AX["trechos"]):
        serie(lin_tr + 2 * i, "%s - Q requerida" % t,
              lambda k, i=i: "=%s" % _aux(AX, "QREQ_TRECHO", i, gcl(3 + k)))
        serie(lin_tr + 2 * i + 1, "%s - Q disponivel" % t,
              lambda k, i=i: "=%s" % _aux(AX, "QDISP_TRECHO", i, gcl(3 + k)))
        linhas_tr.append(lin_tr + 2 * i)

    def montar(tipo, titulo, y, linhas, ancora, altura=9, largura=30):
        ch = LineChart() if tipo == "linha" else BarChart()
        if tipo != "linha":
            ch.type = "col"
        ch.title = titulo
        ch.y_axis.title = y
        ch.x_axis.title = "Semana da safra"
        ch.height, ch.width = altura, largura
        cats = Reference(ws, min_col=3, max_col=2 + nsem, min_row=base)
        for ln in linhas:
            ref = Reference(ws, min_col=2, max_col=2 + nsem, min_row=ln)
            ch.append(Series(ref, title_from_data=True))
        ch.set_categories(cats)
        for s in ch.series:
            s.smooth = False
        ws.add_chart(ch, ancora)

    montar("linha", "GRAFICO 1 - Demanda hidrica da RDM (mm/semana)", "mm/semana",
           [base + 1, base + 2, base + 3, base + 4], "B5")
    montar("barra", "GRAFICO 2 - Volume semanal (m3)", "m3/semana",
           [base + 6, base + 7, base + 8], "B24")
    montar("barra", "GRAFICO 3 - Demanda semanal por casa de bomba (m3)", "m3/semana",
           list(range(lin_casa, lin_casa + len(AX["casas"]))), "B43", altura=10)
    ws2 = wb.create_sheet("GRAFICO_TRECHOS")
    est.cabecalho_pagina(ws2, "GRAFICO 4 - Trechos de canal: vazao requerida x disponivel",
                         "Uma comparacao por trecho. Onde a linha de Q requerida ultrapassa a de Q disponivel, "
                         "o trecho e o gargalo daquela semana (ver RESUMO_TRECHOS, bloco de deficit).", "N")
    linha_anc = 5
    for i, t in enumerate(AX["trechos"]):
        ch = LineChart()
        ch.title = "Trecho %s - Q requerida x Q disponivel (m3/h)" % t
        ch.y_axis.title = "m3/h"
        ch.x_axis.title = "Semana da safra"
        ch.height, ch.width = 8, 30
        for ln in (lin_tr + 2 * i, lin_tr + 2 * i + 1):
            ch.append(Series(Reference(ws, min_col=2, max_col=2 + nsem, min_row=ln), title_from_data=True))
        ch.set_categories(Reference(ws, min_col=3, max_col=2 + nsem, min_row=base))
        ws2.add_chart(ch, "B%d" % linha_anc)
        linha_anc += 17


# ========================================================= MEMORIA_CALCULO
FORMULAS_DOC = [
    ("1. Evapotranspiracao de referencia", "ETo", "Entrada do modelo (mm/dia)",
     "CLIMA_ETo_CHUVA coluna G. Nao e calculada aqui: vem da serie climatica informada. "
     "ETo da semana = soma dos 7 dias (SEMANAS coluna H)."),
    ("2. Coeficiente de cultura", "Kc = f(DAE / duracao do ciclo)",
     "Interpolacao linear entre 4 pontos",
     "Fracao do ciclo = DAE / duracao. Fase inicial: Kc = Kc_ini. Desenvolvimento: interpolacao linear "
     "de Kc_ini ate Kc_med. Fase media: Kc = Kc_med. Fase final: interpolacao linear de Kc_med ate Kc_fim. "
     "Curvas em PARAMETROS_CULTURAS, por cultura e grupo de ciclo."),
    ("3. Evapotranspiracao da cultura", "ETc = ETo x Kc", "mm/semana",
     "PLANEJAMENTO_SEMANAL coluna T. Diaria, semanal, mensal e acumulada saem por SOMASES sobre a mesma coluna."),
    ("4. Chuva util", "Chuva_util = MIN(Chuva_bruta x FATOR_CHUVA_UTIL ; LIMITE_CHUVA_UTIL_DIA)", "mm/dia",
     "CLIMA_ETo_CHUVA coluna L. O usuario pode digitar a chuva util diretamente (coluna M), que entao prevalece. "
     "Cenario deterministico: sem Monte Carlo, sem sorteio de cenarios."),
    ("5. Necessidade liquida de irrigacao", "NL = MAX(0 ; ETc - Chuva_util)", "mm/semana",
     "PLANEJAMENTO_SEMANAL coluna W. ETc, chuva util e NL ficam em colunas separadas e visiveis."),
    ("6. Lamina bruta", "LB = NL / Eficiencia de aplicacao", "mm/semana",
     "PLANEJAMENTO_SEMANAL coluna Y. A eficiencia vem de CADASTRO_PIVOS (por pivo) e, na falta dela, "
     "de EFICIENCIA_PADRAO. DADO PENDENTE enquanto nao houver eficiencia medida."),
    ("7. Volume", "Volume = LB x Area x 10", "m3",
     "1 mm em 1 ha = 10 m3 (parametro MM_PARA_M3_HA). PLANEJAMENTO_SEMANAL coluna Z. "
     "Agregado por pivo, casa, trecho, semana, mes e safra por SOMASES."),
    ("8. Vazao do pivo", "Q_pivo = Area x Lamina nominal x 10 / HORAS_OPERACAO_DIA", "m3/h",
     "CADASTRO_PIVOS coluna I. Mesma regra do arquivo oficial da RDM. Se houver vazao de projeto cadastrada "
     "(coluna H), ela prevalece."),
    ("9. Horas necessarias por pivo", "Horas = Volume / Q_pivo", "h/semana",
     "PLANEJAMENTO_SEMANAL coluna AB."),
    ("10. Horas disponiveis", "Horas_disp = HORAS_OPERACAO_DIA x dias da semana", "h/semana",
     "SEMANAS coluna K. Com 21 h/dia e 7 dias: 147 h/semana. O 21 e o 7 sao parametros, "
     "nunca numeros dentro de formula."),
    ("11. Utilizacao do pivo", "Utilizacao = Horas necessarias / Horas disponiveis", "%",
     "PLANEJAMENTO_SEMANAL coluna AD. Faixas: ate LIMITE_CONFORTAVEL = OK; ate LIMITE_ATENCAO = ATENCAO; "
     "ate LIMITE_CRITICO = CRITICO; acima = INVIAVEL."),
    ("12. Demanda da casa de bomba", "Volume_casa = soma dos volumes dos pivos ligados a ela", "m3/semana",
     "AUX_MATRIZES bloco VOL_CASA. O vinculo pivo-casa vem de CADASTRO_PIVOS, nunca de formula fixa."),
    ("13. Vazao requerida da casa", "Q_req_casa = Volume_casa / Horas disponiveis na semana", "m3/h",
     "AUX_MATRIZES bloco QREQ_CASA."),
    ("14. Utilizacao da casa", "Utilizacao_casa = Q_req_casa / Q_operacional_casa", "%",
     "AUX_MATRIZES bloco UTIL_CASA. Deficit = Q_req - Q_operacional."),
    ("15. Demanda do trecho de canal",
     "Volume_trecho = SOMA(matriz_trecho_casa x Volume_casa)", "m3/semana",
     "AUX_MATRIZES bloco VOL_TRECHO. A matriz TRECHO x CASA (CADASTRO_TRECHOS_CANAL) define quais casas "
     "dependem de cada trecho; em canal em serie, o trecho a montante soma toda a demanda a jusante."),
    ("16. Vazao requerida do trecho", "Q_req_trecho = Volume_trecho / Horas disponiveis na semana", "m3/h",
     "AUX_MATRIZES bloco QREQ_TRECHO. Com 21 h/dia x 7 dias, o divisor e 147 h."),
    ("17. Utilizacao e deficit do trecho",
     "Utilizacao = Q_req / Q_disp   |   Deficit = Q_req - Q_disp", "% e m3/h",
     "Deficit <= 0: SEM DEFICIT. Deficit > 0: TRECHO HIDRAULICAMENTE LIMITANTE."),
    ("18. Capacidade do sistema em 21 h/dia",
     "Capacidade_semana = Q_operacional_total x HORAS_OPERACAO_DIA x dias", "m3/semana",
     "CAPACIDADE_21H coluna F. Deficit de volume, de vazao e de horas saem da mesma comparacao."),
    ("19. Horas e vazao adicionais",
     "Horas_adicionais = Volume / Q_total - Horas disponiveis   |   "
     "Q_adicional = Volume / Horas disponiveis - Q_total", "h e m3/h",
     "CAPACIDADE_21H colunas K e L. Respondem 'quanto falta' quando o status e INVIAVEL."),
]


def aba_memoria(wb):
    ws = wb.create_sheet("MEMORIA_CALCULO")
    est.cabecalho_pagina(ws, "MEMORIA DE CALCULO",
                         "Cada formula do modelo, onde ela mora e o que ela significa. "
                         "Hierarquia: CLIMA -> ETo -> Kc -> ETc -> chuva util -> necessidade -> volume -> horas -> "
                         "casa de bomba -> trecho de canal -> capacidade -> gargalo.", "F")
    escrever_tabela(ws, ["Etapa", "Formula", "Unidade", "Onde esta e como funciona"],
                    [40, 62, 20, 110])
    for i, (etapa, formula, unidade, onde) in enumerate(FORMULAS_DOC):
        r = DAT + i
        ws.cell(row=r, column=2, value=etapa).style = "rotulo"
        c = ws.cell(row=r, column=3, value=formula)
        c.style = "texto"
        c.alignment = est.Alignment(vertical="center", wrap_text=True)
        ws.cell(row=r, column=4, value=unidade).style = "texto"
        c = ws.cell(row=r, column=5, value=onde)
        c.style = "texto"
        c.alignment = est.Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 46
    nota(ws, DAT + len(FORMULAS_DOC) + 1,
         "A pasta nao usa VBA nem macro. Tudo e SOMASES, INDICE/CORRESP, SE, MAXIMO, MINIMO, SOMARPRODUTO, "
         "validacao de dados, formatacao condicional e graficos nativos.", 5)


# ===================================================== PENDENCIAS_CADASTRO
def aba_pendencias(wb, D):
    ws = wb.create_sheet("PENDENCIAS_CADASTRO")
    est.cabecalho_pagina(ws, "PENDENCIAS DE CADASTRO  -  o que falta para o modelo virar compromisso",
                         "Nada aqui foi inventado para fechar conta. Cada linha aponta um dado ausente nos "
                         "arquivos oficiais, o que o modelo esta usando no lugar dele e o efeito no resultado.",
                         "H")
    escrever_tabela(ws, ["Variavel", "Equipamento / abrangencia", "Dado necessario", "Unidade",
                         "O que o modelo faz enquanto isso", "Impacto no calculo", "Onde preencher"],
                    [34, 30, 52, 14, 62, 62, 34])
    fixas = [
        ("Evapotranspiracao de referencia (ETo)", "Toda a RDM",
         "Serie diaria de ETo da safra (estacao, INMET ou banco climatico)", "mm/dia",
         "Usa a curva mensal de referencia PROVISORIA de config, marcada em vermelho na aba de clima.",
         "Escala TODA a demanda. Erro de 10% em ETo e erro de ~10% em volume, horas e utilizacao.",
         "CLIMA_ETo_CHUVA, coluna 'ETo usuario'"),
        ("Chuva e chuva util", "Toda a RDM",
         "Serie de chuva da safra e criterio de aproveitamento adotado pela fazenda", "mm/dia e %",
         "Chuva de referencia PROVISORIA x FATOR_CHUVA_UTIL, limitada por LIMITE_CHUVA_UTIL_DIA.",
         "Define quanto da ETc e abatido antes da irrigacao. Subestimar a chuva util superdimensiona o sistema.",
         "CLIMA_ETo_CHUVA, colunas de chuva / PARAMETROS_GERAIS"),
        ("Eficiencia de aplicacao", "Todos os pivos",
         "Eficiencia medida por pivo (CUC/CUD, ensaio de uniformidade)", "fracao",
         "EFICIENCIA_PADRAO provisoria, replicada em todos os pivos.",
         "Divide a necessidade liquida. Entra direto na lamina bruta, no volume e nas horas.",
         "CADASTRO_PIVOS, coluna 'Eficiencia' / PARAMETROS_GERAIS"),
        ("Vazao de projeto dos pivos", "Todos os pivos",
         "Vazao nominal de cada pivo, do projeto ou do ensaio de campo", "m3/h",
         "Calcula Area x Lamina nominal x 10 / horas de operacao, reproduzindo a regra do arquivo oficial. "
         "A lamina nominal e a maior lamina/dia ja registrada no arquivo oficial de cada pivo.",
         "Define as horas necessarias de cada pivo e a utilizacao dentro das 21 h/dia.",
         "CADASTRO_PIVOS, coluna 'Vazao cadastrada'"),
        ("Vazao operacional das casas de bomba", "RM01 a RM07",
         "Vazao de recalque efetiva de cada casa (conjunto motobomba em operacao)", "m3/h",
         "Usa a soma das vazoes dos pivos ligados a casa, o que faz a utilizacao tender a 100% por construcao.",
         "Sem ela nao existe folga nem deficit real por casa de bomba: o diagnostico fica cego nesse nivel.",
         "CADASTRO_CASAS_BOMBA, coluna 'Vazao operacional cadastrada'"),
        ("Capacidade hidraulica dos trechos de canal", "TR01 a TR07",
         "Vazao maxima e vazao operacional de cada trecho", "m3/h",
         "Repete a soma das casas atendidas: utilizacao 100% e deficit zero por construcao.",
         "E o nivel em que costuma estar o gargalo real. Sem esse dado o modelo nao consegue apontar "
         "trecho limitante.", "CADASTRO_TRECHOS_CANAL, colunas de vazao"),
        ("Topologia do canal", "Rede de canais da RDM",
         "Quais casas de bomba dependem de cada trecho (montante/jusante)", "matriz 1/0",
         "Matriz provisoria em diagonal: 1 trecho por casa, sem acumulo a jusante.",
         "Em canal em serie, o trecho a montante carrega a soma de todas as casas a jusante. "
         "Com a diagonal, a demanda acumulada esta subestimada.",
         "CADASTRO_TRECHOS_CANAL, matriz TRECHO x CASA"),
        ("Curvas de Kc", "Todas as culturas",
         "Kc calibrado para as cultivares e o manejo da RDM", "adimensional",
         "Curvas FAO-56 de partida, editaveis, com 4 fases e interpolacao linear.",
         "Multiplica a ETo. Junto com a ETo, e o que define a ETc de cada semana.",
         "PARAMETROS_CULTURAS"),
    ]
    linha = DAT
    for f in fixas:
        for j, v in enumerate(f):
            c = ws.cell(row=linha, column=2 + j, value=v)
            c.style = "rotulo" if j == 0 else "texto"
            c.alignment = est.Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[linha].height = 58
        linha += 1
    if D["pendencias"]:
        ws.cell(row=linha + 1, column=2, value="INCONSISTENCIAS ENCONTRADAS NA IMPORTACAO DA ROTACAO").style = "secao"
        linha += 2
        for j, h in enumerate(["Variavel", "Equipamento", "Dado encontrado", "Unidade", "Impacto no calculo"]):
            ws.cell(row=linha, column=2 + j, value=h).style = "cabecalho"
        linha += 1
        for p in D["pendencias"]:
            for j, v in enumerate((p["variavel"], p["equipamento"], p["dado"], p["unidade"], p["impacto"])):
                c = ws.cell(row=linha, column=2 + j, value=v)
                c.style = "rotulo" if j == 0 else "texto"
                c.alignment = est.Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[linha].height = 30
            linha += 1
    return linha


# ================================================================= LEIA-ME
def aba_leiame(wb, D):
    ws = wb.create_sheet("LEIA-ME")
    est.cabecalho_pagina(ws, "SIMULADOR DE PLANEJAMENTO HIDRICO DA SAFRA  -  UNIDADE RDM",
                         "Modelo de capacidade hidrica e operacional. Filtrado exclusivamente para a RDM: "
                         "Karitel e demais unidades nao entram em nenhum calculo.", "H")
    linha = 5
    blocos = [
        ("O QUE ESTA PASTA RESPONDE", [
            "Quanto cada cultura, pivo, casa de bomba e trecho de canal demanda de agua, semana a semana.",
            "Quantas horas cada pivo precisa funcionar e se cabe nas horas disponiveis por dia.",
            "Em que semana esta o pico, e qual pivo, qual casa de bomba e qual trecho e o gargalo.",
            "Quanto da demanda a chuva util cobre e quanto sobra para a irrigacao.",
            "Volume total de agua da safra, em mm e em m3.",
        ]),
        ("ONDE MEXER (celulas amarelas, texto azul)", [
            "PARAMETROS_GERAIS: horas por dia, eficiencia padrao, fator de chuva util, faixas de alerta.",
            "CADASTRO_PIVOS: area, lamina nominal, vazao de projeto e eficiencia de cada pivo.",
            "CADASTRO_CASAS_BOMBA: vazao operacional de recalque de cada casa.",
            "CADASTRO_TRECHOS_CANAL: capacidade dos trechos e a matriz TRECHO x CASA (topologia do canal).",
            "PARAMETROS_CULTURAS: curva de Kc por cultura e grupo de ciclo.",
            "ROTACAO_RDM: cultura, cultivar, data de plantio e data final de cada ciclo.",
            "CLIMA_ETo_CHUVA: ETo e chuva, dia a dia.",
        ]),
        ("ONDE LER O RESULTADO", [
            "PAINEL_GESTAO: o coracao do modelo de gestao de agua - em quais casas e dias vai faltar agua, "
            "quanto falta e o que fazer (escalonar plantio, ajustar captacao, priorizar pivos).",
            "BALANCO_DIARIO: o balanco dia a dia do reservatorio de cada casa que alimenta o painel.",
            "DASHBOARD_RDM: os indicadores executivos da safra e a resposta direta sobre as 21 h/dia.",
            "BALANCO_CASAS: por casa de bomba, entrada (captacao), consumo (pivos) e nivel do "
            "reservatorio semana a semana - mostra onde a captacao nao acompanha o consumo.",
            "GRAFICO_CASAS: o mesmo balanco em grafico, uma casa por vez.",
            "CAPACIDADE_21H: semana a semana, nos quatro niveis (pivo, casa, trecho, sistema).",
            "TOP10_SEMANAS: as dez semanas mais criticas da safra, ordenadas automaticamente.",
            "RESUMO_CASAS e RESUMO_TRECHOS: matrizes semana a semana, com indicador selecionavel.",
            "MAPA_CALOR: onde o sistema aperta, em cores.",
            "GRAFICOS e GRAFICO_TRECHOS: as quatro visoes graficas da safra.",
            "BALANCO_CULTURA e BALANCO_PIVO: o fechamento da safra por cultura e por pivo.",
            "PLANEJAMENTO_SEMANAL: a tabela-motor, uma linha por SEMANA x PIVO.",
        ]),
        ("ANTES DE USAR PARA DECIDIR", [
            "Leia a aba PENDENCIAS_CADASTRO. ETo, chuva util, eficiencia, vazao operacional das casas e "
            "capacidade dos trechos NAO constam nos arquivos oficiais.",
            "Enquanto esses campos estiverem provisorios, os numeros indicam ordem de grandeza e o "
            "comportamento relativo entre semanas, pivos e casas, nao um compromisso operacional.",
            "A memoria de calculo de cada formula esta em MEMORIA_CALCULO.",
        ]),
        ("LEGENDA DE CORES", [
            "Fundo amarelo com texto azul: entrada do usuario. E o unico lugar onde se digita.",
            "Texto preto: resultado de formula. Texto verde: valor trazido de outra aba.",
            "Faixa vermelha: DADO PENDENTE. Verde/amarelo/laranja/vermelho nas matrizes: folga, atencao, "
            "limite e capacidade excedida.",
        ]),
    ]
    for titulo, itens in blocos:
        ws.cell(row=linha, column=2, value=titulo).style = "secao"
        linha += 1
        for it in itens:
            c = ws.cell(row=linha, column=2, value="-  " + it)
            c.style = "texto"
            c.alignment = est.Alignment(vertical="top", wrap_text=True, indent=1)
            ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=8)
            ws.row_dimensions[linha].height = 26
            linha += 1
        linha += 1
    ws.column_dimensions["B"].width = 60
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 20
    nota(ws, linha, "Fontes: arquivo oficial de distribuicao de agua da RDM (casas de bomba, pivos, areas, "
                    "pocos, reservacao, lamina praticada) e planilha de rotacao de culturas bienal 26-28 "
                    "(cultura, cultivar, plantio e colheita), filtrada em FAZENDA = RDM e tipo Irrigado.", 8)


ORDEM_ABAS = ["LEIA-ME", "DASHBOARD_RDM", "PAINEL_GESTAO", "CAPACIDADE_21H", "TOP10_SEMANAS", "RESUMO_CASAS",
              "BALANCO_CASAS", "BALANCO_DIARIO", "GRAFICO_CASAS", "RESUMO_TRECHOS", "MAPA_CALOR", "GRAFICOS", "GRAFICO_TRECHOS", "BALANCO_CULTURA",
              "BALANCO_PIVO", "PLANEJAMENTO_SEMANAL", "CADASTRO_PIVOS", "CADASTRO_CASAS_BOMBA",
              "CADASTRO_TRECHOS_CANAL", "PARAMETROS_CULTURAS", "ROTACAO_RDM", "CLIMA_ETo_CHUVA",
              "SEMANAS", "PARAMETROS_GERAIS", "AUX_MATRIZES", "MEMORIA_CALCULO", "PENDENCIAS_CADASTRO"]


def main(caminho_json, saida):
    D = preparar(caminho_json)
    wb = Workbook()
    wb.remove(wb.active)
    est.registrar(wb)

    aba_parametros(wb, D)
    RP = aba_pivos(wb, D)
    RC = aba_casas(wb, D, RP)
    RT = aba_trechos(wb, D, RC)
    RCL = aba_clima(wb, D)
    RPC = aba_culturas(wb, D, RCL)
    RR = aba_rotacao(wb, D, RP)
    RS = aba_semanas(wb, D, RCL)
    RPS = aba_planejamento(wb, D, RP, RR, RPC, RS)
    AX = aba_aux(wb, D, RPS, RS, RC, RT)
    aba_resumo_casas(wb, D, AX)
    BC = aba_balanco_casas(wb, D, AX, RC, RS)
    aba_grafico_casas(wb, D, AX, BC)
    BD = aba_balanco_diario(wb, D, AX, RC, RCL)
    aba_painel_gestao(wb, D, AX, RC, BD)
    aba_resumo_trechos(wb, D, AX)
    aba_mapa_calor(wb, D, AX)
    RCP = aba_capacidade(wb, D, AX, RS)
    aba_top10(wb, D, AX, RCP)
    aba_balanco_cultura(wb, D, RPS, RR)
    aba_balanco_pivo(wb, D, RPS, RR, RP)
    aba_dashboard(wb, D, AX, RPS, RP, RCP, BC)
    aba_graficos(wb, D, AX, RS)
    aba_memoria(wb)
    aba_pendencias(wb, D)
    aba_leiame(wb, D)

    wb._sheets.sort(key=lambda s: ORDEM_ABAS.index(s.title)
                    if s.title in ORDEM_ABAS else len(ORDEM_ABAS))
    wb.active = 0
    wb.save(saida)
    print("abas=%d pivos=%d ciclos=%d semanas=%d linhas_planejamento=%d pendencias=%d"
          % (len(wb.sheetnames), len(D["pivos"]), len(D["ciclos"]), len(D["semanas"]),
             RPS["fim"] - RPS["ini"] + 1, len(D["pendencias"])))


# ======================================================== BALANCO_CASAS
def aba_balanco_casas(wb, D, AX, RC, RS):
    """Balanco de agua por casa de bomba, no estilo da planilha original:
    entrada (captacao dos pocos) x saida/consumo (demanda dos pivos) x nivel
    do reservatorio semana a semana."""
    ws = wb.create_sheet("BALANCO_CASAS")
    est.cabecalho_pagina(ws, "BALANCO DE AGUA POR CASA DE BOMBA",
                         "Para cada casa: ENTRADA = captacao dos pocos, SAIDA/CONSUMO = demanda dos pivos ligados "
                         "a ela, e a projecao do NIVEL DO RESERVATORIO. Premissa (igual a original): alimentacao "
                         "constante e reservatorio comeca cheio.", "L")
    nsem = AX["nsem"]
    casas = AX["casas"]
    ncasa = len(casas)
    ci = RC["ini"]                       # linha da primeira casa em CADASTRO_CASAS_BOMBA

    def entrada(i):
        # volume de captacao/dia (arquivo oficial) x dias da semana
        return "CADASTRO_CASAS_BOMBA!$J$%d*DIAS_POR_SEMANA" % (ci + i)

    def saida(i, k):
        return _aux(AX, "VOL_CASA", i, gcl(3 + k))

    def cap(i):
        return "CADASTRO_CASAS_BOMBA!$O$%d" % (ci + i)

    ws.column_dimensions["B"].width = 30
    blocos = [
        ("ENTRADA - Volume de captacao dos pocos (m3/semana)", "entrada", "num0"),
        ("SAIDA / CONSUMO - Demanda dos pivos da casa (m3/semana)", "saida", "num0"),
        ("SALDO da semana = entrada - consumo (m3/semana)", "saldo", "num0"),
        ("NIVEL DO RESERVATORIO ao fim da semana (m3) - comeca cheio", "nivel", "num0"),
        ("STATUS do abastecimento da casa", "status", "texto"),
    ]
    linha = 5
    anchor_nivel = None
    pos_saldo = None
    for titulo, chave, fmt in blocos:
        ws.cell(row=linha, column=2, value=titulo).style = "secao"
        ws.cell(row=linha + 1, column=2, value="Casa de bomba").style = "cabecalho"
        for k in range(nsem):
            ws.cell(row=linha + 1, column=3 + k, value="=SEMANAS!$B$%d" % (RS["ini"] + k)).style = "cabecalho"
            ws.column_dimensions[gcl(3 + k)].width = 9
        ini = linha + 2
        if chave == "nivel":
            anchor_nivel = ini
        if chave == "saldo":
            pos_saldo = ini
        for i, casa in enumerate(casas):
            r = ini + i
            ws.cell(row=r, column=2, value=casa).style = "rotulo"
            for k in range(nsem):
                C = gcl(3 + k)
                if chave == "entrada":
                    v = "=%s" % entrada(i)
                elif chave == "saida":
                    v = "=%s" % saida(i, k)
                elif chave == "saldo":
                    v = "=%s-%s" % (entrada(i), saida(i, k))
                elif chave == "nivel":
                    prev = cap(i) if k == 0 else "%s%d" % (gcl(2 + k), r)
                    # reservatorio: comeca cheio, soma o saldo, limitado entre 0 e a capacidade
                    v = "=MIN(%s,MAX(0,%s+%s%d))" % (cap(i), prev, gcl(3 + k), pos_saldo + i)
                else:
                    niv = "%s%d" % (C, anchor_nivel + i)
                    sal = "%s%d" % (C, pos_saldo + i)
                    v = ('=IF(%s<=0,"RESERVATORIO ZERADO - captacao insuficiente",'
                         'IF(%s<0,"CONSUMINDO RESERVA","OK"))' % (niv, sal))
                ws.cell(row=r, column=3 + k, value=v).style = fmt
        faixa = "C%d:%s%d" % (ini, gcl(2 + nsem), ini + ncasa - 1)
        if chave == "saldo":
            ws.conditional_formatting.add(faixa, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))
        elif chave == "nivel":
            ws.conditional_formatting.add(faixa, ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                end_type="max", end_color="C6EFCE"))
        elif chave == "status":
            for txt, cor in (("OK", est.VERDE_OK), ("CONSUMINDO", "FFEB9C"),
                             ("RESERVATORIO", est.VERMELHO_PEND)):
                ws.conditional_formatting.add(faixa, CellIsRule(
                    operator="beginsWith", formula=['"%s"' % txt],
                    fill=est.PatternFill("solid", fgColor=cor)))
        linha = ini + ncasa + 2

    # ---- resumo por casa
    ws.cell(row=linha, column=2, value="RESUMO DA SAFRA POR CASA DE BOMBA").style = "secao"
    linha += 1
    cabs = ["Casa de bomba", "Entrada total (m3)", "Consumo total (m3)", "Saldo da safra (m3)",
            "Capacidade do reservatorio (m3)", "Consumo medio (m3/semana)",
            "Entrada media (m3/semana)", "Semanas consumindo reserva", "Semanas com reservatorio zerado",
            "Situacao"]
    larg = [16, 16, 16, 16, 18, 17, 17, 15, 16, 40]
    for j, (h, w) in enumerate(zip(cabs, larg)):
        ws.cell(row=linha, column=2 + j, value=h).style = "cabecalho"
        ws.column_dimensions[gcl(2 + j)].width = w
    ent_ini = 7
    sai_ini = ent_ini + ncasa + 4
    sal_ini = sai_ini + ncasa + 4
    niv_ini = sal_ini + ncasa + 4
    sta_ini = niv_ini + ncasa + 4
    fim_col = gcl(2 + nsem)
    for i, casa in enumerate(casas):
        r = linha + 1 + i
        ws.cell(row=r, column=2, value=casa).style = "rotulo"
        ws.cell(row=r, column=3, value="=SUM(C%d:%s%d)" % (ent_ini + i, fim_col, ent_ini + i)).style = "num0"
        ws.cell(row=r, column=4, value="=SUM(C%d:%s%d)" % (sai_ini + i, fim_col, sai_ini + i)).style = "num0"
        ws.cell(row=r, column=5, value="=$C%d-$D%d" % (r, r)).style = "num0"
        ws.cell(row=r, column=6, value="=%s" % cap(i)).style = "num0"
        ws.cell(row=r, column=7, value="=IFERROR($D%d/N_SEMANAS,0)" % r).style = "num0"
        ws.cell(row=r, column=8, value="=IFERROR($C%d/N_SEMANAS,0)" % r).style = "num0"
        ws.cell(row=r, column=9, value="=COUNTIF(C%d:%s%d,\"<0\")" % (sal_ini + i, fim_col, sal_ini + i)
                ).style = "num0"
        ws.cell(row=r, column=10, value="=COUNTIF(C%d:%s%d,0)" % (niv_ini + i, fim_col, niv_ini + i)
                ).style = "num0"
        ws.cell(row=r, column=11,
                value='=IF($J%d>0,"CAPTACAO INSUFICIENTE - reservatorio zera em "&$J%d&" semana(s)",'
                      'IF($E%d>=0,"OK - captacao cobre o consumo da safra",'
                      '"ATENCAO - consumo da safra supera a captacao, sustentado pela reserva"))'
                      % (r, r, r)).style = "texto"
    fim = linha + ncasa
    nota(ws, fim + 2,
         "ENTRADA = Volume de captacao/dia da casa (arquivo oficial RDM) x dias por semana. CONSUMO = demanda "
         "de todos os pivos ligados a casa. O nivel do reservatorio comeca cheio, soma o saldo de cada semana e "
         "fica limitado entre zero e a capacidade cadastrada. 'Reservatorio zerado' = a captacao mais a reserva "
         "nao cobrem o consumo daquela semana.", 11)
    return {"ent_ini": ent_ini, "sai_ini": sai_ini, "niv_ini": niv_ini,
            "resumo_ini": linha + 1, "resumo_fim": linha + ncasa}


# =================================================== GRAFICO_BALANCO_CASAS
def aba_grafico_casas(wb, D, AX, BC):
    ws = wb.create_sheet("GRAFICO_CASAS")
    est.cabecalho_pagina(ws, "GRAFICO - Balanco de agua por casa de bomba",
                         "Uma casa por grafico: ENTRADA (captacao) x CONSUMO (demanda dos pivos) x NIVEL do "
                         "reservatorio, semana a semana. Onde o consumo supera a entrada, o reservatorio cai.",
                         "N")
    nsem = AX["nsem"]
    origem = wb["BALANCO_CASAS"]
    linha_anc = 5
    for i, casa in enumerate(AX["casas"]):
        ch = LineChart()
        ch.title = "%s - Entrada x Consumo x Nivel do reservatorio" % casa
        ch.y_axis.title = "m3"
        ch.x_axis.title = "Semana da safra"
        ch.height, ch.width = 8, 30
        cats = Reference(origem, min_col=3, max_col=2 + nsem, min_row=BC["ent_ini"] - 1)
        for base in (BC["ent_ini"], BC["sai_ini"], BC["niv_ini"]):
            ref = Reference(origem, min_col=2, max_col=2 + nsem, min_row=base + i)
            ch.append(Series(ref, title_from_data=True))
        ch.set_categories(cats)
        ws.add_chart(ch, "B%d" % linha_anc)
        linha_anc += 17

# ========================================================= BALANCO_DIARIO
def aba_balanco_diario(wb, D, AX, RC, RCL):
    """Coracao do modelo de gestao de agua: balanco DIARIO do reservatorio de
    cada casa de bomba. Captacao (18 h) entra, consumo dos pivos (21 h) sai, e o
    nivel do reservatorio e projetado dia a dia. E aqui que se enxerga o dia exato
    em que a agua falta - o que a media semanal esconde."""
    ws = wb.create_sheet("BALANCO_DIARIO")
    est.cabecalho_pagina(ws, "BALANCO DIARIO DO RESERVATORIO POR CASA DE BOMBA",
                         "Captacao (entrada, 18 h/dia) x consumo dos pivos (saida) x nivel do reservatorio, dia a "
                         "dia. Reservatorio comeca cheio. Celula de deficit em vermelho = a captacao mais a reserva "
                         "nao cobriram o consumo daquele dia: FALTA DE AGUA.", "X")
    casas = AX["casas"]
    ncasa = len(casas)
    ci = RC["ini"]
    nsem = AX["nsem"]
    ult = gcl(2 + nsem)
    dias = AX["nsem"] * 7

    def capt(i):
        return "CADASTRO_CASAS_BOMBA!$I$%d*HORAS_CAPTACAO_DIA" % (ci + i)

    def cap(i):
        return "CADASTRO_CASAS_BOMBA!$O$%d" % (ci + i)

    def dem_semana(i, semcell):
        vol_row = AX["VOL_CASA"]["ini"] + i
        return "INDEX(AUX_MATRIZES!$C$%d:$%s$%d,MIN(N_SEMANAS,MAX(1,%s)))" % (vol_row, ult, vol_row, semcell)

    # cabecalho: 3 colunas por casa (demanda, nivel, deficit)
    cab1, cab2, dat = 5, 6, 7
    ws.cell(row=cab1, column=2, value="Data").style = "cabecalho"
    ws.cell(row=cab2, column=2, value="").style = "cabecalho"
    ws.cell(row=cab1, column=3, value="Sem").style = "cabecalho"
    ws.cell(row=cab2, column=3, value="").style = "cabecalho"
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 6
    for i, casa in enumerate(casas):
        base = 4 + 3 * i
        ws.merge_cells(start_row=cab1, start_column=base, end_row=cab1, end_column=base + 2)
        c = ws.cell(row=cab1, column=base,
                    value='=%s&"  (capt "&TEXT(%s,"#,##0")&" m3/d - cap "&TEXT(%s,"#,##0")&" m3)"'
                          % ("CADASTRO_CASAS_BOMBA!$B$%d" % (ci + i), capt(i), cap(i)))
        c.style = "cabecalho"
        for j, t in enumerate(("Consumo m3/d", "Nivel m3", "Deficit m3/d")):
            ws.cell(row=cab2, column=base + j, value=t).style = "cabecalho"
            ws.column_dimensions[gcl(base + j)].width = 12
    ws.row_dimensions[cab1].height = 26
    ws.row_dimensions[cab2].height = 24

    for t in range(dias):
        r = dat + t
        ws.cell(row=r, column=2, value="=DATA_INICIO_SAFRA+%d" % t).style = "data"
        ws.cell(row=r, column=3, value="=MIN(N_SEMANAS,MAX(1,INT(($B%d-DATA_INICIO_SAFRA)/DIAS_POR_SEMANA)+1))" % r
                ).style = "num0"
        for i in range(ncasa):
            base = 4 + 3 * i
            dcol, ncol, fcol = gcl(base), gcl(base + 1), gcl(base + 2)
            dem = "%s/DIAS_POR_SEMANA" % dem_semana(i, "$C%d" % r)
            ws.cell(row=r, column=base, value="=%s" % dem).style = "num0"
            prev = cap(i) if t == 0 else "%s%d" % (ncol, r - 1)
            ws.cell(row=r, column=base + 1,
                    value="=MIN(%s,MAX(0,%s+%s-%s%d))" % (cap(i), prev, capt(i), dcol, r)).style = "num0"
            ws.cell(row=r, column=base + 2,
                    value="=MAX(0,%s%d-%s-%s)" % (dcol, r, capt(i), prev)).style = "num0"
    fim = dat + dias - 1
    for i in range(ncasa):
        base = 4 + 3 * i
        ncol, fcol = gcl(base + 1), gcl(base + 2)
        ws.conditional_formatting.add("%s%d:%s%d" % (ncol, dat, ncol, fim), ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max", end_color="C6EFCE"))
        ws.conditional_formatting.add("%s%d:%s%d" % (fcol, dat, fcol, fim), CellIsRule(
            operator="greaterThan", formula=["0"],
            fill=est.PatternFill("solid", fgColor=est.VERMELHO_PEND)))
    ws.freeze_panes = "D%d" % dat
    return {"dat": dat, "fim": fim, "ncasa": ncasa}


# ========================================================== PAINEL_GESTAO
def aba_painel_gestao(wb, D, AX, RC, BD):
    """Onde e quando vai faltar agua - e o que fazer. Le o balanco diario e
    transforma em decisao: dias de deficit por casa, primeiro dia critico, volume
    faltante, horas extras de captacao necessarias e recomendacao."""
    ws = wb.create_sheet("PAINEL_GESTAO")
    est.cabecalho_pagina(ws, "PAINEL DE GESTAO DE AGUA  -  onde e quando vai faltar",
                         "Diagnostico por casa de bomba, a partir do balanco diario do reservatorio. Responde: em "
                         "quais casas e dias a captacao nao acompanha o consumo, quanto falta e o que fazer.", "N")
    casas = AX["casas"]
    ci = RC["ini"]
    dat, fim = BD["dat"], BD["fim"]

    escrever_tabela(ws, ["Casa de bomba", "Captacao (m3/dia)", "Consumo de pico (m3/dia)",
                         "Dias com falta de agua", "Primeiro dia de falta", "Ultimo dia de falta",
                         "Volume total faltante (m3)", "Maior falta em um dia (m3)",
                         "Horas extras de captacao no pico (h/dia)", "Vazao adicional necessaria (m3/h)",
                         "Situacao", "Recomendacao"],
                    [14, 15, 17, 15, 15, 15, 16, 16, 18, 17, 20, 60])
    for i, casa in enumerate(casas):
        r = DAT + i
        base = 4 + 3 * i
        dcol, fcol = gcl(base), gcl(base + 2)
        rng_def = "BALANCO_DIARIO!$%s$%d:$%s$%d" % (fcol, dat, fcol, fim)
        rng_dem = "BALANCO_DIARIO!$%s$%d:$%s$%d" % (dcol, dat, dcol, fim)
        rng_dat = "BALANCO_DIARIO!$B$%d:$B$%d" % (dat, fim)
        ws.cell(row=r, column=2, value="=CADASTRO_CASAS_BOMBA!$B$%d" % (ci + i)).style = "rotulo"
        ws.cell(row=r, column=3, value="=CADASTRO_CASAS_BOMBA!$I$%d*HORAS_CAPTACAO_DIA" % (ci + i)).style = "num0"
        ws.cell(row=r, column=4, value="=MAX(%s)" % rng_dem).style = "num0"
        ws.cell(row=r, column=5, value='=COUNTIF(%s,">0")' % rng_def).style = "num0"
        ws.cell(row=r, column=6,
                value='=IFERROR(INDEX(%s,MATCH(1,INDEX(--(%s>0),0),0)),"-")' % (rng_dat, rng_def)).style = "data"
        ws.cell(row=r, column=7,
                value='=IFERROR(LOOKUP(2,1/(%s>0),%s),"-")' % (rng_def, rng_dat)).style = "data"
        ws.cell(row=r, column=8, value="=SUM(%s)" % rng_def).style = "num0"
        ws.cell(row=r, column=9, value="=MAX(%s)" % rng_def).style = "num0"
        ws.cell(row=r, column=10,
                value="=IFERROR($I%d/CADASTRO_CASAS_BOMBA!$I$%d,0)" % (r, ci + i)).style = "num1"
        ws.cell(row=r, column=11, value="=IFERROR($I%d/HORAS_CAPTACAO_DIA,0)" % r).style = "num0"
        ws.cell(row=r, column=12,
                value='=IF($E%d=0,"OK - captacao cobre o consumo",'
                      'IF($E%d<=7,"ATENCAO - falta pontual","CRITICO - falta recorrente"))' % (r, r)).style = "texto"
        ws.cell(row=r, column=13,
                value='=IF($E%d=0,"Sem acao: a captacao atende o consumo o ano todo.",'
                      '"Falta em "&$E%d&" dia(s), pico de "&TEXT($I%d,"#,##0")&" m3/dia (~"&TEXT($J%d,"0.0")&'
                      '" h extras de captacao ou +"&TEXT($K%d,"#,##0")&" m3/h). Alternativas: escalonar plantio '
                      'para afastar o pico, reduzir simultaneidade nesta casa, ampliar horas/vazao de captacao ou '
                      'priorizar os pivos criticos.")' % (r, r, r, r, r)).style = "texto"
        ws.row_dimensions[r].height = 42
    fim_p = DAT + len(casas) - 1
    for txt, cor in (("OK", est.VERDE_OK), ("ATENCAO", "FFEB9C"), ("CRITICO", est.VERMELHO_PEND)):
        ws.conditional_formatting.add("L%d:L%d" % (DAT, fim_p), CellIsRule(
            operator="beginsWith", formula=['"%s"' % txt],
            fill=est.PatternFill("solid", fgColor=cor)))
    linha = fim_p + 2
    ws.cell(row=linha, column=2, value="RESUMO DA RDM").style = "secao"
    for j, (rot, form, fmt) in enumerate((
            ("Casas com falta de agua", '=COUNTIF($E$%d:$E$%d,">0")' % (DAT, fim_p), "num0"),
            ("Total de dias-casa com falta", "=SUM($E$%d:$E$%d)" % (DAT, fim_p), "num0"),
            ("Volume total faltante na safra (m3)", "=SUM($H$%d:$H$%d)" % (DAT, fim_p), "num0"),
            ("Casa mais critica",
             '=IFERROR(INDEX($B$%d:$B$%d,MATCH(MAX($E$%d:$E$%d),$E$%d:$E$%d,0)),"-")'
             % (DAT, fim_p, DAT, fim_p, DAT, fim_p), "texto"))):
        ws.cell(row=linha + 1 + j, column=2, value=rot).style = "rotulo"
        ws.cell(row=linha + 1 + j, column=6, value=form).style = fmt
    nota(ws, linha + 6,
         "Como usar para gerir a safra: se uma casa aparece CRITICO, teste no modelo mover a data de plantio dos "
         "pivos dela (aba ROTACAO_RDM) para afastar o pico de consumo, ou ajuste as horas de captacao "
         "(PARAMETROS_GERAIS). Todos os numeros deste painel se refazem na hora.", 13)




if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
