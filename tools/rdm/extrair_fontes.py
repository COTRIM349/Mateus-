# -*- coding: utf-8 -*-
"""Extrai, dos dois arquivos oficiais, o dataset canonico da RDM.

Fontes:
  1) SANTA COLOMBA_RDM_ANALISE DISTRIBUICAO AGUA ... .xlsm  -> hidraulica
  2) COPIA DE ROTACAO DE CULTURAS BIENAL 26-28 ... .xlsx    -> rotacao agricola

Saida: rdm_dataset.json (consumido por construir_planilha.py)
"""
import json
import sys
import unicodedata
from datetime import datetime

import openpyxl


def norm(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.upper()


def d(x):
    return x.strftime("%Y-%m-%d") if isinstance(x, datetime) else None


def extrair_hidraulica(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    # --- Parametros: casa de bomba x pivo x area x reservacao -------------
    ws = wb["Parametros"] if "Parametros" in wb.sheetnames else wb["Parâmetros"]
    linhas = list(ws.iter_rows(min_row=6, max_row=37, max_col=13, values_only=True))
    pivos, reservacao, casa_atual = [], {}, None
    for r in linhas:
        casa, pivo, area = r[0], r[1], r[2]
        if not casa or norm(casa) == "TOTAL":
            continue
        casa = str(casa).strip()
        casa_atual = casa
        pivos.append({"pivo": "PV %s" % str(pivo).strip(), "casa": casa, "area_ha": float(area)})
        # reservacao aparece na linha do ultimo pivo do bloco de cada casa
        res, canal, total = r[3], r[4], r[5]
        for chave, val in (("reservatorio_m3", res), ("canal_m3", canal), ("capacidade_total_m3", total)):
            if isinstance(val, (int, float)):
                reservacao.setdefault(casa_atual, {})[chave] = float(val)

    # --- Pocos de captacao (vazao de entrada) -----------------------------
    pocos = []
    for r in ws.iter_rows(min_row=4, max_row=16, min_col=11, max_col=13, values_only=True):
        if r[0]:
            pocos.append({"poco": str(r[0]).strip(), "id": str(r[1]).strip(),
                          "vazao_m3h": float(r[2])})

    # --- Dashboard: blocos de 19 linhas, um por casa de bomba -------------
    dsh = wb["Dashboard"]
    linhas_d = {i: r for i, r in enumerate(
        dsh.iter_rows(min_row=1, max_row=140, max_col=12, values_only=True), 1)}
    casas = []
    for k in range(7):
        o = k * 19
        casa = str(linhas_d[2 + o][1]).strip()
        pocos_casa, vazoes, turnos = [], [], []
        for j in range(4, 11):
            if linhas_d[7 + o][j]:
                pocos_casa.append(str(linhas_d[7 + o][j]).strip())
                vazoes.append(linhas_d[9 + o][j])
                turnos.append(linhas_d[10 + o][j])
        casas.append({
            "casa": casa,
            "pocos": pocos_casa,
            "vazao_captacao_m3h": float(sum(v for v in vazoes if v)),
            "turno_captacao_h": float(max(t for t in turnos if t)) if turnos else None,
            "volume_captacao_m3_dia": float(linhas_d[11 + o][2] or 0),
            "turno_rega_h": float(linhas_d[13 + o][2] or 0),
            "reservatorio_m3": linhas_d[14 + o][4],
            "canal_m3": linhas_d[14 + o][6],
            "capacidade_total_m3": linhas_d[14 + o][9],
        })

    # --- Lamina nominal por pivo (maximo praticado no arquivo oficial) ----
    base = wb["data_Base"]
    bloco = list(base.iter_rows(min_row=2, max_row=583, min_col=21, max_col=52, values_only=True))
    cabec = bloco[0]
    lamina = {}
    for j, h in enumerate(cabec):
        if not h:
            continue
        vals = [r[j] for r in bloco[3:] if isinstance(r[j], (int, float)) and r[j] > 0]
        if vals:
            pivo = "PV %s" % h.split("PV_")[1]
            lamina[pivo] = {"lamina_max_mm_dia": max(vals),
                            "lamina_media_mm_dia": round(sum(vals) / len(vals), 2),
                            "dias_com_lamina": len(vals)}
    for p in pivos:
        p.update(lamina.get(p["pivo"], {}))

    # --- Janela de simulacao do arquivo oficial ---------------------------
    pl = wb["Planejamento_Irrigacao"]
    janela = {"data_inicio": d(pl["C6"].value), "data_fim": d(pl["C7"].value)}
    wb.close()
    return {"pivos": pivos, "casas": casas, "pocos": pocos,
            "reservacao": reservacao, "janela_arquivo_oficial": janela}


def extrair_rotacao(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["Rotação"]
    ciclos = []
    for i, r in enumerate(ws.iter_rows(min_row=10, max_row=200, max_col=19, values_only=True), 10):
        if norm(r[0]) != "RDM" or norm(r[2]) != "IRRIGADO":
            continue
        talhao = str(r[1]).strip()
        if not talhao.isdigit():
            continue                      # 201..204 e ASV2 sao sequeiro/nao-pivo
        base = {"talhao": talhao, "pivo": "PV %s" % talhao, "area_rotacao_ha": r[3]}
        # cultura anterior: so a data de encerramento e conhecida
        if r[4]:
            ciclos.append(dict(base, ordem="C0", rotulo="ANTERIOR", cultura=str(r[4]).strip(),
                               cultivar=(str(r[5]).strip() if r[5] else None),
                               plantio=None, fim=d(r[6])))
        for ordem, rotulo, c in (("C1", "CICLO 1", (7, 8, 9, 10)),
                                 ("C2", "CICLO 2", (11, 12, 13, 14)),
                                 ("C3", "CICLO 3", (15, 16, 17, 18))):
            cult, var, ini, fim = (r[c[0]], r[c[1]], r[c[2]], r[c[3]])
            if not (cult or var or ini or fim):
                continue
            ciclos.append(dict(base, ordem=ordem, rotulo=rotulo,
                               cultura=(str(cult).strip() if cult else None),
                               cultivar=(str(var).strip() if var else None),
                               plantio=d(ini), fim=d(fim)))

    # --- Duracao de ciclo por cultivar (aba de validacao) -----------------
    v = wb["Base de dados validação"]
    linhas = list(v.iter_rows(min_row=4, max_row=35, max_col=28, values_only=True))
    cultivares = []
    for r in linhas:                       # SOJA: cultivar col B, janela col D
        if r[1]:
            cultivares.append({"cultura": "SOJA", "cultivar": str(r[1]).strip(),
                               "grm": r[2], "duracao_dias": r[3]})
        if r[4]:                           # ALGODAO: cultivar col E, ciclo F, janela G
            cultivares.append({"cultura": "ALGODAO", "cultivar": str(r[4]).strip(),
                               "grm": (str(r[5]).strip() if r[5] else None), "duracao_dias": r[6]})
        if r[12]:                          # MILHO_GRAO: cultivar col M, ciclo N, janela O
            cultivares.append({"cultura": "MILHO_GRAO", "cultivar": str(r[12]).strip(),
                               "grm": (str(r[13]).strip() if r[13] else None), "duracao_dias": r[14]})
    wb.close()
    return {"ciclos": ciclos, "cultivares": cultivares}


if __name__ == "__main__":
    hid = extrair_hidraulica(sys.argv[1])
    rot = extrair_rotacao(sys.argv[2])
    dados = dict(hid, **rot)
    with open(sys.argv[3], "w", encoding="utf-8") as fh:
        json.dump(dados, fh, ensure_ascii=False, indent=1)
    print("pivos=%d casas=%d pocos=%d ciclos=%d cultivares=%d"
          % (len(dados["pivos"]), len(dados["casas"]), len(dados["pocos"]),
             len(dados["ciclos"]), len(dados["cultivares"])))
