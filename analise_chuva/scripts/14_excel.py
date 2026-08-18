# -*- coding: utf-8 -*-
"""Planilha consolidada com todas as tabelas da analise."""
import pandas as pd, numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
BASE = Path(__file__).resolve().parent.parent
D, S = BASE/"dados", BASE/"saidas"
OUT = Path("/home/user/Mateus-/analise_chuva/Impacto_da_Chuva_Soja_Algodao_2526.xlsx")

AZ, CZ, BR = "1F3B4D", "E8EEF2", "F5F7F8"
hf = Font(bold=True, color="FFFFFF", size=10); hp = PatternFill("solid", fgColor=AZ)
tf = Font(bold=True, size=13, color=AZ)
bd = Border(*[Side("thin", color="D0D8DE")]*4)

def aba(wb, nome, titulo, nota, df, larguras=None, pct=(), dec=()):
    ws = wb.create_sheet(nome[:31])
    ws["A1"] = titulo; ws["A1"].font = tf
    ws["A2"] = nota;   ws["A2"].font = Font(size=9, italic=True, color="55666F")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(df.columns),6))
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 30
    r0 = 4
    for j, c in enumerate(df.columns, 1):
        cel = ws.cell(r0, j, str(c)); cel.font = hf; cel.fill = hp
        cel.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[r0].height = 34
    for i, (_, row) in enumerate(df.iterrows(), r0+1):
        for j, c in enumerate(df.columns, 1):
            v = row[c]
            if isinstance(v, (pd.Timestamp,)): v = v.date() if pd.notna(v) else None
            if pd.isna(v): v = None
            cel = ws.cell(i, j, v); cel.border = bd; cel.font = Font(size=10)
            if i % 2 == 0: cel.fill = PatternFill("solid", fgColor=BR)
            if isinstance(v, float): cel.number_format = "0.0" if c in dec else "#,##0.0"
    for j, c in enumerate(df.columns, 1):
        w = (larguras or {}).get(c, max(11, min(30, int(df[c].astype(str).str.len().max() or 10)+3)))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(r0+1, 1)
    return ws

wb = Workbook(); wb.remove(wb.active)

# ---------------- capa
ws = wb.create_sheet("Resumo")
ws["A1"] = "Impacto da chuva no plantio da soja e do algodão — safra 2025/26"
ws["A1"].font = Font(bold=True, size=16, color=AZ)
ws["A2"] = "Fazendas Karitel (KRT) e Rio do Meio (RDM) · Cocos/BA · 8.305 ha de algodão, 7.037 ha de soja"
ws["A2"].font = Font(size=11, color="55666F")
lin = [
 ("", ""),
 ("REGRA APLICADA", "Dia com chuva > 5 mm = não agricultável, mais R dias residuais. R calibrado nos registros reais de plantio."),
 ("", ""),
 ("Dias da campanha do algodão (02/12/25 a 19/03/26)", 108),
 ("Dias com chuva > 5 mm (KRT / RDM)", "33 / 33"),
 ("Dias bloqueados por efeito residual (R=2)", "29 / 35"),
 ("TOTAL de dias não agricultáveis", "62 (57%) / 68 (63%)"),
 ("Dias agricultáveis restantes", "46 / 40"),
 ("Maior sequência de dias bloqueados", "10 dias / 16 dias"),
 ("", ""),
 ("Dias de máquina realmente necessários", "24 (KRT) / 15 (RDM)"),
 ("Atraso do algodão atribuível à chuva (média pond.)", "2,5 a 7,8 dias"),
 ("Atraso total do algodão (chuva + sucessão)", "4,9 a 10,2 dias"),
 ("Escalonamento deliberado da soja, não explicado por chuva", "94 dias (KRT) / 64 dias (RDM)"),
 ("", ""),
 ("CONCLUSÃO", "A chuva bloqueou a maioria dos dias da campanha, mas o calendário do algodão "
  "foi definido principalmente pela liberação escalonada dos pivôs pela soja. A capacidade de "
  "máquina não foi o gargalo."),
]
for i, (a, b) in enumerate(lin, 4):
    ws.cell(i, 1, a).font = Font(bold=a in ("REGRA APLICADA","CONCLUSÃO") or a.startswith("TOTAL"), size=10)
    c = ws.cell(i, 2, b); c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 52; ws.column_dimensions["B"].width = 75

# ---------------- abas
T = pd.read_csv(S/"TALHAO_CONSOLIDADO.csv")
aba(wb, "Talhoes", "Métricas por talhão (pivô) — algodão 25/26",
    "Uma linha por evento de plantio. 'camp_' = janela comum da campanha (02/12 a 19/03), permite comparar talhões. "
    "'jan_' = janela própria do talhão (liberação até o plantio). API com k=0,85.", T)

R = pd.read_csv(S/"cenarios_resumo.csv")
aba(wb, "Cenarios", "Cenários contrafactuais A–E",
    "H1 = pivôs sem soja 25/26 registrada presos à data real (limite INFERIOR do efeito da chuva). "
    "H2 = esses pivôs livres desde 02/12 (limite SUPERIOR). A verdade está entre os dois.", R)

C = pd.read_csv(S/"cenarios_por_talhao.csv")
aba(wb, "Cenarios_por_talhao", "Data simulada de plantio por talhão em cada cenário",
    "Datas de término do plantio de cada pivô sob cada cenário e hipótese de liberação.", C)

for arq, nome, tit, nota in [
  ("calib_residual.csv","Calib_residual","Calibração do efeito residual",
   "Dias secos (≤5 mm) classificados pela distância ao último evento >5 mm. P(operar) medida nos registros reais "
   "de plantio. Base = solo já recuperado. D+1 é significativo (Fisher p=0,0035)."),
  ("calib_chuva_dia.csv","Calib_chuva_dia","Probabilidade de operar x chuva do dia",
   "Note que o dia da chuva ainda registra operação (chuva de tarde, plantio de manhã). O bloqueio real recai sobre D+1."),
  ("calib_api_k.csv","Calib_API","Escolha do decaimento k do API",
   "AUC de separação entre dias com e sem operação. Ver limitação: o API contínuo separa mal (AUC≈0,55, p=0,15)."),
  ("decomposicao_dias.csv","Decomposicao_dias","Decomposição dos 108 dias da campanha",
   "Cada dia da campanha classificado por fazenda: bloqueado pela chuva, sem área liberada, ou agricultável."),
  ("serie_diaria_api.csv","Serie_diaria","Série diária de chuva, API e bloqueio",
   "Chuva mediana da fazenda, API (k=0,85) e classificação de cada dia."),
  ("produtividade_algodao.csv","Produtividade","Produtividade x época de plantio",
   "ATENÇÃO: só 21 dos 68 eventos (30% da área) têm produtividade fechada, todos plantados entre 02/12 e 11/01. "
   "Não extrapolar para fevereiro/março."),
  ("aud_chuva_vs_relatorio.csv","Aud_fonte_auxiliar","Auditoria: pluviômetros x relatório de irrigação",
   "Confronto do acumulado por pivô entre as duas fontes independentes. Viés mediano +1,1%; 49 de 54 dentro de ±15%."),
  ("aud_series.csv","Aud_series","Auditoria por série de pluviômetro",
   "Acumulado, contagem de zeros e zeros suspeitos (0 mm no sensor com ≥10 mm na mediana da fazenda)."),
  ("aud_pares.csv","Aud_pares","Auditoria: sensores redundantes no mesmo pivô",
   "Razão entre o sensor secundário e o primário. Razão ~0,3–0,45 indica perda de eventos — série reprovada."),
]:
    p = S/arq
    if p.exists():
        df = pd.read_csv(p)
        aba(wb, nome, tit, nota, df)

# ---------------- metodo e limitacoes
ws = wb.create_sheet("Metodo_e_limitacoes")
ws["A1"] = "Método, hipóteses e limitações"; ws["A1"].font = Font(bold=True, size=15, color=AZ)
txt = [
 ("FONTES", ""),
 ("", "CHUVAS_RDM.xlsx e chuva_karitel.xlsx — chuva diária por pluviômetro (39 e 79 séries)."),
 ("", "relatorio_soja.pdf (Karitel) e soja_safra_rdm.pdf (Rio do Meio) — plantio, colheita, área e produtividade da soja por pivô."),
 ("", "Produtividade_e_Epoca_de_Plantio_Algodao_safra_2526.xlsx — plantio, emergência, área, variedade e produtividade do algodão."),
 ("AUDITORIA", ""),
 ("", "Dias ausentes: 42% (KRT) e 49% (RDM) do calendário não têm linha no arquivo. Verificado que nenhum dia presente tem chuva zero em todos os sensores: o arquivo só exporta dias com chuva. Dia ausente foi tratado como 0 mm."),
 ("", "Validação independente: o acumulado por pivô bate com o relatório de irrigação (viés mediano +1,1%; 49 de 54 pivôs dentro de ±15%). Isso confirma que a série não tem perda relevante de eventos."),
 ("", "Duplicados: nenhum par (data, pluviômetro) repetido. Unidades: mín 0, máx 157 mm/dia, acumulados de 340 a 1.390 mm — compatível com mm/dia no Oeste da Bahia."),
 ("", "Séries reprovadas: KRT-004B e KRT-006B (sensores secundários com 30% e 43% do acumulado do primário no mesmo pivô e zeros em dias de chuva geral). Excluídas."),
 ("", "Outliers mantidos: os maiores valores (157 mm em 21/11) ocorreram em dias de chuva regional confirmada por vários sensores. Nenhum outlier isolado (>80 mm com mediana da fazenda <2 mm)."),
 ("CALIBRAÇÃO", ""),
 ("", "R (dias residuais) foi calibrado nos registros reais de plantio, não arbitrado. P(operar) = 0,12 em D+1 e 0,25 em D+2, contra 0,52 na base (solo recuperado). Em D+3 já volta à base."),
 ("", "Teste exato de Fisher: D+1 vs base p=0,0035 (significativo); D+2 p=0,079; D+3 p=1,0. D+1 e D+2 juntos: p=0,004, queda de 65% na chance de operar."),
 ("", "Sensibilidade: R=1 (recuperação rápida) → 48/54 dias bloqueados; R=2 (calibrado) → 62/68; R=3 (lenta) → 73/80."),
 ("", "O API contínuo separou mal os dias operados (AUC 0,55, p=0,15 por permutação). O bloqueio é curto e quase binário, não uma função suave da umidade acumulada. O API é reportado como descritor, não como regra."),
 ("HIPÓTESES", ""),
 ("", "Capacidade operacional = p90 do ha/dia observado: 250 ha/dia (KRT) e 150 ha/dia (RDM)."),
 ("", "A data de plantio registrada representa o pivô plantado naquele dia (não há registro de plantio parcial)."),
 ("", "Liberação do talhão = data de colheita da soja, para os 25 pivôs (3.187 ha) com sucessão soja→algodão documentada."),
 ("", "Para os 43 pivôs restantes (5.118 ha) a cultura antecessora não consta nas fontes. Por isso os cenários foram rodados sob duas hipóteses (H1 e H2) que definem um intervalo, em vez de um número único."),
 ("", "Maturidade da soja = plantio + P25 do ciclo da fazenda (117–118 dias), tomado como colheita sem espera."),
 ("LIMITAÇÕES", ""),
 ("", "Produtividade do algodão disponível para apenas 21 dos 68 eventos (30% da área), todos plantados entre 02/12 e 11/01. A colheita dos plantios de fevereiro e março ainda não fechou. Não é possível medir nesta safra o efeito do plantio tardio sobre a produtividade além de 11/01."),
 ("", "A associação observada (-2,9 @/ha por dia de atraso, controlando fazenda) vale apenas dentro dessa janela de 40 dias e não é prova de causalidade: variedade, solo e manejo variam junto com a data."),
 ("", "Não há registro de horas de máquina, de plantio parcial nem de motivo de parada. Dias sem plantio registrado em condição agricultável não puderam ser atribuídos a uma causa."),
 ("", "Não há dados de estação meteorológica externa nem de umidade de solo por talhão para a janela de plantio; a validação cruzada usou o relatório de irrigação como fonte auxiliar."),
 ("", "A cultura antecessora dos 43 pivôs sem soja registrada é a maior fonte de incerteza do estudo."),
]
r = 3
for a, b in txt:
    if a:
        ws.cell(r, 1, a).font = Font(bold=True, size=11, color=AZ); r += 1
    if b:
        c = ws.cell(r, 2, "• " + b); c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[r].height = 30; r += 1
ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 130

wb.save(OUT); print("gravado:", OUT)
import subprocess; print(subprocess.run(["ls","-la",str(OUT)],capture_output=True,text=True).stdout)
