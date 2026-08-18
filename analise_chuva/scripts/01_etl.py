# -*- coding: utf-8 -*-
"""ETL: consolida chuva diaria (KRT/RDM), soja (PDF) e algodao (XLSX) em CSVs limpos."""
import re, unicodedata, datetime as dt
from pathlib import Path
import openpyxl, pdfplumber, pandas as pd

UP = Path("/root/.claude/uploads/38efc676-a594-577a-8efe-82677373c579")
OUT = Path("/home/user/Mateus-/analise_chuva/dados")
OUT.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------- CHUVA
def norm_pluvio(col, fazenda):
    c = str(col).strip()
    m = re.match(r'(?i)^pluvio_(?:pv)?0*(\d+)$', c)          # serie primaria (pluvio_pvNN / pluvio_NN)
    if m:
        return f"{fazenda}-{int(m.group(1)):03d}", "pivo"
    m = re.match(r'(?i)^pluvio\s+0*(\d+)$', c)                # serie secundaria "Pluvio NN" (com espaco)
    if m:
        return f"{fazenda}-{int(m.group(1)):03d}B", "pivo_secundario"
    m = re.match(r'(?i)^pluvio\s+sc\s+([0-9a-f]+)$', c)
    if m:
        return f"{fazenda}-SC{m.group(1).upper()}", "estacao"
    return f"{fazenda}-{c}", "outro"

def le_chuva(arquivo, fazenda):
    wb = openpyxl.load_workbook(UP / arquivo, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    regs, totais = [], {}
    for r in rows[1:]:
        d = r[0]
        if isinstance(d, dt.datetime):
            data = d.date()
            for j in range(2, len(hdr)):          # pula col 'Media'
                if hdr[j] is None: continue
                pid, tipo = norm_pluvio(hdr[j], fazenda)
                regs.append((fazenda, data, pid, tipo, r[j]))
        elif isinstance(d, str) and d.strip().lower() == "total":
            for j in range(2, len(hdr)):
                if hdr[j] is None: continue
                pid, _ = norm_pluvio(hdr[j], fazenda)
                totais[pid] = r[j]
    df = pd.DataFrame(regs, columns=["fazenda", "data", "pluvio_id", "tipo", "mm_bruto"])
    tot = pd.DataFrame([(k, v) for k, v in totais.items()],
                       columns=["pluvio_id", "total_declarado_mm"])
    return df, tot

ch_rdm, tot_rdm = le_chuva("10808bcf-CHUVAS_RDM.xlsx", "RDM")
ch_krt, tot_krt = le_chuva("2fd44531-chuva_karitel.xlsx", "KRT")
chuva = pd.concat([ch_rdm, ch_krt], ignore_index=True)
totais = pd.concat([tot_rdm, tot_krt], ignore_index=True)
chuva["mm"] = pd.to_numeric(chuva["mm_bruto"], errors="coerce")
chuva.to_csv(OUT / "chuva_long_bruto.csv", index=False)
totais.to_csv(OUT / "chuva_totais_declarados.csv", index=False)
print("chuva bruta:", chuva.shape, "| pluviometros:", chuva.pluvio_id.nunique())

# ---------------------------------------------------------------- SOJA (PDF)
CAMPOS = r'(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s+' + r'([\d.,]+)\s+' * 7 + r'([\d.,]+)\s*$'
LINHA = re.compile(r'^Piv[oô]\s+(\S+?)(?:\s*-\s*ABC)?\s*-\s*Soja\s*25/26\s+' + CAMPOS)
LABEL = re.compile(r'^Piv[oô]\s+(\S+?)(?:\s*-\s*ABC)?\s*-\s*Soja\s*$')   # label quebrado em 2 linhas
DADOS = re.compile(r'^' + CAMPOS)

def num(s):
    return float(s.replace(".", "").replace(",", ".")) if s else None

def monta(fazenda, pivo, g):
    return dict(fazenda=fazenda, pivo=pivo,
                plantio=dt.datetime.strptime(g[0], "%d/%m/%Y").date(),
                colheita=dt.datetime.strptime(g[1], "%d/%m/%Y").date(),
                dias_manejados=int(g[2]), area_ha=num(g[3]),
                prod_sc_ha=num(g[4]), chuva_mm=num(g[5]), etc_mm=num(g[6]),
                etpc_mm=num(g[7]), stress_pct=num(g[8]),
                irrig_mm=num(g[9]), excesso_mm=num(g[10]))

def le_soja(arquivo, fazenda):
    regs = []
    with pdfplumber.open(UP / arquivo) as pdf:
        for pg in pdf.pages[:6]:
            pend = None
            for ln in (pg.extract_text() or "").split("\n"):
                ln = " ".join(ln.split())
                m = LINHA.match(ln)
                if m:
                    regs.append(monta(fazenda, m.group(1), m.groups()[1:])); pend = None
                    continue
                ml = LABEL.match(ln)
                if ml:
                    pend = ml.group(1); continue
                md = DADOS.match(ln)
                if md and pend:
                    regs.append(monta(fazenda, pend, md.groups())); pend = None
    return pd.DataFrame(regs)

soja = pd.concat([le_soja("cd71bc36-relatorio_soja.pdf", "KRT"),
                  le_soja("d0954938-soja_safra_rdm.pdf", "RDM")], ignore_index=True)
soja = soja.drop_duplicates(subset=["fazenda", "pivo"])
soja["pivo_id"] = soja.fazenda + "-" + soja.pivo.str.extract(r'(\d+)')[0].astype(int).map("{:03d}".format)
soja.to_csv(OUT / "soja.csv", index=False)
print("soja:", soja.shape, "| area:", soja.area_ha.sum())

# ---------------------------------------------------------------- ALGODAO
wb = openpyxl.load_workbook(UP / "869fd5d6-Produtividade_e_E_poca_de_Plantio_Algoda_o_safra_2526.xlsx",
                            data_only=True)
ws = wb.active
regs, ult = [], None
for r in ws.iter_rows(min_row=3, values_only=True):
    faz, pv, area, var, pl, em, kg, arr = r[:8]
    if faz is None or str(faz).strip() in ("Total", ""): break
    faz = str(faz).strip().upper()
    pivo = int(re.sub(r'\D', '', str(pv)))
    if pl is not None:
        ult = (pl, em)
    plantio, emerg = (pl, em) if pl is not None else (ult if ult else (None, None))
    regs.append(dict(fazenda=faz, pivo=pivo, pivo_id=f"{faz}-{pivo:03d}",
                     area_ha=area, variedade=str(var).strip() if var else None,
                     plantio=plantio.date() if plantio else None,
                     emergencia=emerg.date() if emerg else None,
                     kg_totais=kg, arrobas_ha=arr,
                     data_propria=pl is not None))
alg = pd.DataFrame(regs)
alg.to_csv(OUT / "algodao_lotes.csv", index=False)

# agrega por pivo (talhao operacional): 1 evento de plantio por data
algp = (alg.dropna(subset=["plantio"])
           .groupby(["fazenda", "pivo_id", "pivo", "plantio", "emergencia"], as_index=False)
           .agg(area_ha=("area_ha", "sum"),
                kg_totais=("kg_totais", "sum"),
                variedades=("variedade", lambda s: "; ".join(sorted(set(x for x in s if x))))))
algp["arrobas_ha"] = (algp.kg_totais / 15 / algp.area_ha).where(algp.kg_totais > 0)
algp.to_csv(OUT / "algodao_pivo.csv", index=False)
print("algodao lotes:", alg.shape, "| eventos plantio:", algp.shape,
      "| area:", algp.area_ha.sum(), "| com produtividade:", algp.arrobas_ha.notna().sum())
